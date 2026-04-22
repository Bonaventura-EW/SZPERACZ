#!/usr/bin/env python3
"""
SZPERACZ OLX — Rekonstrukcja historii added/removed per dzień z Excela.

Dla każdego profilu i każdego dnia wylicza:
- added: ID obecne w scanie dnia D, których nie było w ostatnim scanie poprzedniego dnia
- removed: ID obecne w ostatnim scanie dnia D-1, których nie ma w scanie dnia D

Zapisuje wyniki do data/dashboard_data.json (do pól `added` i `removed` w daily_counts).
"""

import os
import json
import logging
from collections import defaultdict
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("rebuild-flows")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "szperacz_olx.xlsx")
JSON_PATH = os.path.join(BASE_DIR, "data", "dashboard_data.json")

# Arkusze profilowe (pomijamy historia_cen, podsumowanie)
PROFILE_SHEETS = {
    "wszystkie_pokoje", "pokojewlublinie", "poqui", "artymiuk",
    "dawny_patron", "mzuri", "villahome",
}


def extract_listing_id(row_vals):
    """
    Zwraca ID ogłoszenia z wiersza.
    W Excelu były dwa formaty:
      - stary (luty/marzec): ID w col 12 (index 11)
      - nowy (kwiecień+):    ID w col 17 (index 16)
    Szukamy pierwszego niepustego pola wyglądającego jak OLX ID.
    """
    # Preferujemy col 17 (nowszy format — bardziej niezawodny)
    for idx in (16, 11):  # 0-indexed: col 17, col 12
        if idx >= len(row_vals):
            continue
        val = row_vals[idx]
        if val is None:
            continue
        sval = str(val).strip()
        if not sval or sval == "—":
            continue
        # ID może być alfanumeryczny (18KAEc) lub numeryczny (1057082751)
        # Odrzucamy URLe i wartości za długie / za krótkie
        if sval.startswith("http"):
            continue
        if len(sval) < 4 or len(sval) > 20:
            continue
        # Musi zawierać tylko alfanum
        if not sval.replace("-", "").isalnum():
            continue
        return sval
    return None


def collect_daily_ids_per_profile(wb):
    """
    Zwraca dict: profile -> date -> set(listing_ids)
    Bierze ID listingów z ostatniego skanu danego dnia (jeśli był >1).
    """
    result = defaultdict(lambda: defaultdict(set))

    for sheet_name in wb.sheetnames:
        if sheet_name not in PROFILE_SHEETS:
            continue

        ws = wb[sheet_name]
        log.info(f"Przetwarzam arkusz: {sheet_name} ({ws.max_row} wierszy)")

        # Zbieramy wszystkie skany: (data, godzina) -> set(ids)
        # następnie dla każdej daty bierzemy skan o najpóźniejszej godzinie
        scan_ids = defaultdict(set)  # (date, time) -> set(ids)
        scan_times_per_date = defaultdict(list)  # date -> [time1, time2, ...]

        current_scan_key = None  # (date, time) aktualnego bloku skanu

        for row in ws.iter_rows(min_row=2, values_only=True):
            # row jest tuple; może mieć mniej niż 17 elementów
            row = list(row) + [None] * (17 - len(row))

            date_val = row[0]
            time_val = row[1]
            count_val = row[2]  # tylko w wierszu summary

            # Normalizacja daty do str YYYY-MM-DD
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, str):
                date_str = date_val.strip()[:10] if date_val.strip() else None
            else:
                date_str = None

            # Normalizacja godziny
            if hasattr(time_val, "strftime"):
                time_str = time_val.strftime("%H:%M")
            elif isinstance(time_val, str):
                time_str = time_val.strip()
            else:
                time_str = None

            if not date_str:
                continue

            # Wiersz summary (zawiera count) otwiera nowy blok skanu
            if count_val is not None and isinstance(count_val, (int, float)):
                current_scan_key = (date_str, time_str or "00:00")
                if time_str and time_str not in scan_times_per_date[date_str]:
                    scan_times_per_date[date_str].append(time_str)
                continue

            # Wiersz listingu — przypisujemy do aktualnego skanu
            if current_scan_key is None:
                # Fallback: użyj daty+godziny z wiersza
                current_scan_key = (date_str, time_str or "00:00")

            lid = extract_listing_id(row)
            if lid:
                scan_ids[current_scan_key].add(lid)

        # Dla każdej daty wybierz skan o najpóźniejszej godzinie
        for date_str in scan_times_per_date:
            times = sorted(scan_times_per_date[date_str])
            latest_time = times[-1] if times else "00:00"
            key = (date_str, latest_time)
            if key in scan_ids:
                result[sheet_name][date_str] = scan_ids[key]
            else:
                # Jeśli nie ma dokładnego dopasowania — weź unię wszystkich skanów tego dnia
                union = set()
                for (d, t), ids in scan_ids.items():
                    if d == date_str:
                        union |= ids
                result[sheet_name][date_str] = union

        # Dodaj też daty które miały tylko jeden skan bez wielokrotnego tracking
        for (d, t), ids in scan_ids.items():
            if d not in result[sheet_name]:
                result[sheet_name][d] = ids

        log.info(f"  {sheet_name}: {len(result[sheet_name])} dni, "
                 f"{sum(len(s) for s in result[sheet_name].values())} listingów (suma)")

    return result


def compute_daily_flows(daily_ids_per_profile, counts_per_profile):
    """
    Zwraca: profile -> date -> {"added": N, "removed": M}
    Dla każdego dnia porównuje ID z ostatnim *kompletnym* poprzednim dniem (chronologicznie).

    Dni uznajemy za "niekompletne" jeśli count > 0 w dashboard, ale w Excelu nie ma
    listingów z ID (bug scrapera w pewnych okresach). Takie dni dostają flow=None.
    """
    result = {}

    for profile, daily_ids in daily_ids_per_profile.items():
        sorted_dates = sorted(daily_ids.keys())
        flows = {}
        counts = counts_per_profile.get(profile, {})

        prev_ids = None
        prev_date = None

        for date in sorted_dates:
            current_ids = daily_ids[date]
            expected_count = counts.get(date)

            # Dzień niekompletny: w Excelu brakuje listingów, ale count > 0
            is_incomplete = (
                len(current_ids) == 0
                and expected_count is not None
                and expected_count > 0
            )

            if is_incomplete:
                # Nie zmieniamy prev_ids — ten dzień zostanie pominięty jako baza
                flows[date] = {"added": None, "removed": None}
                continue

            if prev_ids is None:
                # Pierwszy kompletny dzień — brak bazy do porównania
                flows[date] = {"added": None, "removed": None}
            else:
                added = len(current_ids - prev_ids)
                removed = len(prev_ids - current_ids)
                flows[date] = {"added": added, "removed": removed}

            prev_ids = current_ids
            prev_date = date

        result[profile] = flows
        complete_days = sum(1 for f in flows.values() if f["added"] is not None)
        log.info(f"  {profile}: wyliczono flow dla {len(flows)} dni ({complete_days} z pełnymi danymi)")

    return result


def merge_into_dashboard(flows_per_profile):
    """Wpisuje flowy do dashboard_data.json (pola added/removed w daily_counts[])."""
    if not os.path.exists(JSON_PATH):
        log.error(f"Brak {JSON_PATH}")
        return

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    profiles = data.get("profiles", {})
    updated_count = 0
    missing_count = 0

    for pk, flows in flows_per_profile.items():
        if pk not in profiles:
            log.warning(f"  Profil {pk} nieobecny w dashboard_data.json — pomijam")
            continue

        dc = profiles[pk].get("daily_counts", [])
        for entry in dc:
            date = entry.get("date")
            if date in flows:
                flow = flows[date]
                entry["added"] = flow["added"]
                entry["removed"] = flow["removed"]
                updated_count += 1
            else:
                # Brak danych w Excelu dla tego dnia (nie powinno się zdarzyć)
                entry.setdefault("added", None)
                entry.setdefault("removed", None)
                missing_count += 1

    log.info(f"Zaktualizowano {updated_count} wpisów daily_counts, {missing_count} bez danych z Excela")

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log.info(f"Zapisano: {JSON_PATH}")


def main():
    log.info("=" * 60)
    log.info("REBUILD DAILY FLOWS — added/removed per dzień z Excela")
    log.info("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        log.error(f"Brak {EXCEL_PATH}")
        return

    wb = load_workbook(EXCEL_PATH, read_only=True)

    log.info("\n--- KROK 1: zbieranie ID listingów per profil per dzień ---")
    daily_ids = collect_daily_ids_per_profile(wb)

    log.info("\n--- KROK 2: wczytywanie count'ów z dashboard_data.json (do detekcji niekompletnych dni) ---")
    counts_per_profile = {}
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            dashboard = json.load(f)
        for pk, pdata in dashboard.get("profiles", {}).items():
            counts_per_profile[pk] = {
                entry["date"]: entry.get("count", 0)
                for entry in pdata.get("daily_counts", [])
            }

    log.info("\n--- KROK 3: obliczanie flow (added/removed) ---")
    flows = compute_daily_flows(daily_ids, counts_per_profile)

    log.info("\n--- KROK 4: zapis do dashboard_data.json ---")
    merge_into_dashboard(flows)

    # Raport podsumowujący
    log.info("\n--- PODSUMOWANIE ---")
    for pk, fs in flows.items():
        sample_dates = sorted(fs.keys())[-7:]
        log.info(f"{pk}: {len(fs)} dni, ostatnie 7:")
        for d in sample_dates:
            a = fs[d]["added"]
            r = fs[d]["removed"]
            if a is None:
                log.info(f"    {d}: BRAK DANYCH (niekompletny scan)")
            else:
                log.info(f"    {d}: +{a} -{r}")

    log.info("\n✓ Gotowe")


if __name__ == "__main__":
    main()
