#!/usr/bin/env python3
"""
Jednorazowa migracja: wiersze-podsumowania z `data/szperacz_olx.xlsx`
-> append-only ledger `data/history/daily_summary.ndjson`.

Ledger to WIECZNY, niemutowalny zapis trendu (1 linia = 1 skan jednego profilu),
niezależny od `dashboard_data.json` (które trzyma pełne dane per-ogłoszenie).
Migrujemy TYLKO stały, niezawodny format wierszy-podsumowań (count/crosscheck/change),
więc migracja jest w 100% weryfikowalna (count == liczba wierszy-ogłoszeń w bloku).

Idempotentny: jeśli ledger już istnieje, skrypt odmawia nadpisania (chroni append-only).
"""
import json
import os
import sys
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "data", "szperacz_olx.xlsx")
LEDGER = os.path.join(BASE, "data", "history", "daily_summary.ndjson")
NON_PROFILE_SHEETS = {"historia_cen", "podsumowanie"}


def extract_blocks(ws):
    """Zwraca listę bloków skanu: (date, time, count, change, crosscheck, n_listings)."""
    blocks = []
    cur = None
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, time, count, change, cross, title = row[0], row[1], row[2], row[3], row[4], row[5]
        is_summary = (count is not None and isinstance(count, (int, float))) and (title is None)
        is_listing = title is not None
        if is_summary:
            if cur is not None:
                blocks.append((*cur, n))
            cur = (date, time, int(count), change, cross)
            n = 0
        elif is_listing:
            n += 1
    if cur is not None:
        blocks.append((*cur, n))
    return blocks


def main():
    if os.path.exists(LEDGER):
        sys.exit(f"❌ Ledger już istnieje: {LEDGER}. Migracja przerwana (append-only).")
    if not os.path.exists(XLSX):
        sys.exit(f"❌ Brak {XLSX}")

    os.makedirs(os.path.dirname(LEDGER), exist_ok=True)
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    profiles = [s for s in wb.sheetnames if s not in NON_PROFILE_SHEETS]

    rows = []          # finalne rekordy ledgera
    mismatches = 0
    for sh in profiles:
        for date, time, count, change, cross, n_listings in extract_blocks(wb[sh]):
            if count != n_listings:
                mismatches += 1
                print(f"⚠️  {sh} {date} {time}: count={count} != wiersze={n_listings}")
            rows.append({
                "date": date,
                "time": time,
                "profile": sh,
                "count": count,
                "crosscheck": cross,
                "change": int(change) if isinstance(change, (int, float)) else None,
                "source": "migrated_from_xlsx",
            })

    if mismatches:
        sys.exit(f"❌ {mismatches} niezgodności count!=wiersze — migracja PRZERWANA.")

    # Deterministyczna kolejność: (date, time, profile)
    rows.sort(key=lambda r: (r["date"], r["time"], r["profile"]))

    with open(LEDGER, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    print(f"✅ Zapisano {len(rows)} linii -> {LEDGER}")
    return len(rows)


if __name__ == "__main__":
    main()
