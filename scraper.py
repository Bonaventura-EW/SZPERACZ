#!/usr/bin/env python3
"""
SZPERACZ OLX — Autonomiczny agent monitorujący ogłoszenia OLX.
Scrape'uje profile, śledzi ceny, zapisuje do Excela i generuje JSON dla dashboardu.
"""

import requests
from bs4 import BeautifulSoup
import json
import os
import re
import time
import random
import logging
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ─── Configuration ───────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("szperacz")

PROFILES = {
    "wszystkie_pokoje": {
        "url": "https://www.olx.pl/nieruchomosci/stancje-pokoje/lublin/",
        "label": "Wszystkie pokoje w Lublinie",
        "is_category": True,
    },
    "pokojewlublinie": {
        "url": "https://www.olx.pl/oferty/uzytkownik/3cxbz/",
        "label": "pokojewlublinie",
        "is_category": False,
        "uuid": "23314f02-9f24-4232-afe0-102bda498af4",
    },
    "poqui": {
        "url": "https://www.olx.pl/oferty/uzytkownik/p8eWV/",
        "label": "poqui",
        "is_category": False,
        "uuid": "06418904-e0dc-4842-9f24-5173b0a2c6a9",
    },
    "artymiuk": {
        "url": "https://www.olx.pl/oferty/uzytkownik/BAm3j/",
        "label": "artymiuk",
        "is_category": False,
        "uuid": "3365c8c4-7f43-4ca8-be96-26afcf62f8ed",
    },
    "dawny_patron": {
        "url": "https://www.olx.pl/oferty/uzytkownik/uD2d4/",
        "label": "dawny patron",
        "is_category": False,
        "uuid": "c9941865-a87e-49d9-9d13-920e2a8fead3",
    },
    "mzuri": {
        "url": "https://www.olx.pl/oferty/uzytkownik/4avCO/",
        "label": "mzuri",
        "is_category": False,
        "uuid": "aef14b8b-6252-4f27-adb1-b2b73bd8dea6",
    },
    "villahome": {
        "url": "https://www.olx.pl/oferty/uzytkownik/2MKggM/",
        "label": "villahome",
        "is_category": False,
        "uuid": "6da7ff53-b041-4214-af55-8dc1ea2f30ff",
    },
    "stylowe_pokoje_ania": {
        "url": "https://www.olx.pl/oferty/uzytkownik/1WLoW/",
        "label": "stylowe pokoje-ania",
        "is_category": False,
        "uuid": "6ce22f3f-2dc7-4299-99e6-3c0d7ab9df19",
    },
    "myrent": {
        "url": "https://www.olx.pl/oferty/uzytkownik/56DT9/",
        "label": "MyRent",
        "is_category": False,
        "uuid": "a17ed560-8913-4f97-9c67-0ebb5c2041c0",
    },
    "mat": {
        "url": "https://www.olx.pl/oferty/uzytkownik/4B6oQ/",
        "label": "MAT",
        "is_category": False,
        "uuid": "3e1e5f35-ba20-43bb-87a7-4587cb67af17",
    },
}

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")
API_DIR = os.path.join(DOCS_DIR, "api")
EXCEL_PATH = os.path.join(DATA_DIR, "szperacz_olx.xlsx")
JSON_PATH = os.path.join(DATA_DIR, "dashboard_data.json")
HISTORY_DIR = os.path.join(DATA_DIR, "history")
HISTORY_LEDGER = os.path.join(HISTORY_DIR, "daily_summary.ndjson")
API_STATUS_PATH = os.path.join(API_DIR, "status.json")
API_HISTORY_PATH = os.path.join(API_DIR, "history.json")

# Ogłoszenia z ceną >= tej krotności średniej ceny w danym profilu są odrzucane
# jako błędne dane (literówka w cenie, ogłoszenie nie-pokoju itp.) — patrz filter_price_outliers().
PRICE_OUTLIER_MULTIPLIER = 10

# Skan, który pobrał mniej niż tę część ogłoszeń deklarowanych w nagłówku strony,
# traktujemy jak błąd scrapera (ochrona danych jak przy count==0) — patrz is_header_shortfall().
HEADER_SHORTFALL_RATIO = 0.5

# Progi alertu "masowe zniknięcie ogłoszeń" w docs/api/status.json — patrz generate_api_json().
MASS_REMOVAL_MIN = 10      # min. liczba zniknięć, żeby w ogóle rozważać alert
MASS_REMOVAL_RATIO = 0.3   # zniknięcia >= 30% poprzedniego stanu profilu = anomalia

# Próg alertu "stale_listings": ogłoszenie nieobecne w tylu kolejnych skanach z rzędu
# (missed_scans), a wciąż uznawane za aktywne — może znaczyć, że OLX zmienił komunikat
# o nieaktualności i verify_listing_active() przestała wykrywać martwe ogłoszenia.
# Próg = 12 (a nie 5): zwykła rotacja wyników OLX potrafi ukryć żywe ogłoszenie na kilka
# skanów z rzędu i wygenerować fałszywy alarm (patrz §7 — carried/missed_scans). Dopiero
# ~2 tygodnie nieobecności przy verify=aktywne to sygnał realnego problemu, nie rotacji.
STALE_MISSED_SCANS_MIN = 12


def is_header_shortfall(result):
    """
    True, gdy skan pobrał drastycznie mniej ogłoszeń niż deklaruje nagłówek strony
    (np. 50 z 650 → crosscheck 'best_of_two'). Taki skan traktujemy jak błąd scrapera:
    NIE archiwizujemy, NIE nadpisujemy current_listings, NIE dopisujemy do ledgera.
    Incydent 2026-07-11: taki skan zarchiwizował 595 ogłoszeń, które istniały na OLX.
    """
    header = result.get("header_count")
    count = result.get("count", 0)
    return bool(header and header > 0 and count < header * HEADER_SHORTFALL_RATIO)

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3 Safari/605.1.15",
]


# ─── HTTP Session ────────────────────────────────────────────────────────────

def get_session():
    s = requests.Session()
    ua = random.choice(USER_AGENTS)
    s.headers.update({
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "pl-PL,pl;q=0.9",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    })
    # Add retries with exponential backoff
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    retry_strategy = Retry(
        total=3,
        backoff_factor=2,
        status_forcelist=[429, 500, 502, 503, 504],
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s


def verify_listing_active(url: str, timeout: int = 10) -> bool:
    """
    Sprawdza czy ogłoszenie OLX jest nadal aktywne przez bezpośredni GET na jego URL.
    Zwraca True jeśli ogłoszenie istnieje i jest aktywne, False jeśli zostało usunięte/wygasło.
    W razie błędu sieciowego zwraca True (fail-safe: nie archiwizuj przy wątpliwości).
    """
    # UWAGA: bez gołego "404" — status 404 sprawdzamy po resp.status_code; substring
    # "404" w treści strony trafiał w ID/hashe aktywnych ogłoszeń (fałszywa archiwizacja).
    INACTIVE_PHRASES = [
        "To ogłoszenie jest już nieaktualne",
        "Ogłoszenie nieaktywne",
        "oferta wygasła",
        "oferta została usunięta",
        "Nie znaleźliśmy tej strony",
    ]
    try:
        session = get_session()
        resp = session.get(url.split("?")[0], timeout=timeout, allow_redirects=True)
        # 404 (Not Found) i 410 (Gone) = ogłoszenie usunięte/wygasłe. OLX zwraca 410
        # dla ofert zdjętych przez użytkownika lub po wygaśnięciu — strona 410 nie zawiera
        # żadnej z INACTIVE_PHRASES (tylko generyczny placeholder), więc status HTTP jest
        # jedynym pewnym sygnałem. Potwierdzone 2026-07-24: 5 ogłoszeń stylowe_pokoje_ania
        # zniknęło z listingu profilu, każde dawało HTTP 410, a stara gałąź
        # "status != 200 → aktywne" trzymała je w current_listings bez końca (nie trafiały
        # do archiwum). Patrz §7 CLAUDE.md.
        if resp.status_code in (404, 410):
            return False
        if resp.status_code != 200:
            log.warning(f"[verify] {url[:60]} → HTTP {resp.status_code}, zakładam aktywne")
            return True
        # Frazy szukamy w widocznym tekście strony, nie w surowym HTML — w blobie
        # JSON-a (__PRERENDERED_STATE__) komunikaty szablonu siedzą też na aktywnych stronach.
        text = BeautifulSoup(resp.text, "lxml").get_text(" ").lower()
        for phrase in INACTIVE_PHRASES:
            if phrase.lower() in text:
                log.info(f"[verify] Nieaktywne ('{phrase}'): {url[:60]}")
                return False
        return True
    except Exception as e:
        log.warning(f"[verify] Błąd weryfikacji {url[:60]}: {e} — zakładam aktywne")
        return True


def get_api_session():
    """Session for OLX JSON API endpoints."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "application/json",
        "Accept-Language": "pl-PL,pl;q=0.9",
        "Connection": "keep-alive",
    })
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    adapter = HTTPAdapter(max_retries=Retry(total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504]))
    s.mount("https://", adapter)
    return s


# ─── Parsing Helpers ─────────────────────────────────────────────────────────

def parse_price(text):
    if not text:
        return None
    cleaned = text.replace(" ", "").replace("\xa0", "")
    cleaned = cleaned.split("zł")[0] if "zł" in cleaned else cleaned
    # Odetnij część dziesiętną (grosze) PRZED usunięciem separatorów — inaczej
    # np. "1 260,65 zł" traci przecinek i zamienia się w "126065" (10x realna cena).
    cleaned = re.split(r"[.,]", cleaned)[0]
    cleaned = re.sub(r"[^\d]", "", cleaned)
    if cleaned:
        try:
            return int(cleaned)
        except ValueError:
            return None
    return None


def parse_date_text(text):
    if not text:
        return None, None
    text = text.strip()
    today = datetime.now().strftime("%Y-%m-%d")
    refreshed = None
    published = None

    if "Odświeżono" in text or "odświeżono" in text:
        refreshed = _extract_date(text)
    elif "Dzisiaj" in text or "dzisiaj" in text:
        refreshed = today
        published = today
    else:
        published = _extract_date(text)

    return published, refreshed


def _extract_date(text):
    months_pl = {
        "stycznia": "01", "lutego": "02", "marca": "03", "kwietnia": "04",
        "maja": "05", "czerwca": "06", "lipca": "07", "sierpnia": "08",
        "września": "09", "października": "10", "listopada": "11", "grudnia": "12",
    }
    today = datetime.now()
    if "dzisiaj" in text.lower():
        return today.strftime("%Y-%m-%d")
    if "wczoraj" in text.lower():
        return (today - timedelta(days=1)).strftime("%Y-%m-%d")
    for month_pl, month_num in months_pl.items():
        if month_pl in text.lower():
            match = re.search(r"(\d{1,2})\s+" + month_pl + r"\s+(\d{4})", text.lower())
            if match:
                return f"{match.group(2)}-{month_num}-{match.group(1).zfill(2)}"
            match = re.search(r"(\d{1,2})\s+" + month_pl, text.lower())
            if match:
                return f"{today.year}-{month_num}-{match.group(1).zfill(2)}"
    return None


def promotion_dict_to_fields(promotion: dict) -> dict:
    """
    Convert OLX API 'promotion' dict to is_promoted / promotion_type fields.

    OLX API returns e.g.:
        {"highlighted": false, "top_ad": true, "options": ["promoted_ad_30"], ...}
        {"top_ad": false, "options": ["pushup"], ...}

    Canonical promotion_type values (zgodne z detect_promoted_status):
        'top_ad'    — wyróżnione na górze listy (stała płatna pozycja)
        'highlight' — podświetlone tło
        'urgent'    — pilne
        'premium'   — strona premium
        'pushup'    — jednorazowe podbicie na górę listy (nie stała promocja)
        'unknown'   — inne opcje płatne

    Uwaga: pushup zmienia last_refreshed — to JEST odświeżenie, ale płatne.
    Dlatego is_promoted=True dla pushup, żeby refresh_count go nie zliczał podwójnie.

    Returns:
        {"is_promoted": bool, "promotion_type": str | None}
    """
    if not promotion:
        return {"is_promoted": False, "promotion_type": None}

    opts = promotion.get("options", [])

    # Flagi bezpośrednie mają pierwszeństwo
    if promotion.get("top_ad"):
        return {"is_promoted": True, "promotion_type": "top_ad"}
    if promotion.get("highlighted"):
        return {"is_promoted": True, "promotion_type": "highlight"}
    if promotion.get("urgent"):
        return {"is_promoted": True, "promotion_type": "urgent"}
    if promotion.get("premium_ad_page"):
        return {"is_promoted": True, "promotion_type": "premium"}

    # Opcje z listy options
    if opts:
        opt = opts[0].lower()
        if opt == "pushup":
            return {"is_promoted": True, "promotion_type": "pushup"}
        if opt.startswith("promoted_ad"):
            # promoted_ad_30, promoted_ad_7 itp. → top_ad
            return {"is_promoted": True, "promotion_type": "top_ad"}
        # Inne opcje — zaraportuj jak są
        return {"is_promoted": True, "promotion_type": opt}

    return {"is_promoted": False, "promotion_type": None}


def extract_listing_id(url):
    match = re.search(r"ID([a-zA-Z0-9]+)\.html", url)
    if match:
        return match.group(1)
    parts = url.rstrip("/").split("/")
    return parts[-1] if parts else url


# ─── Card Parsing ────────────────────────────────────────────────────────────

DATE_KEYWORDS = [
    "odświeżono", "dzisiaj", "wczoraj",
    "stycznia", "lutego", "marca", "kwietnia", "maja", "czerwca",
    "lipca", "sierpnia", "września", "października", "listopada", "grudnia",
]


def detect_promoted_status(card):
    """
    Detect if listing is promoted/featured (HTML scraper — kategoria).

    Canonical promotion_type values (zgodne z promotion_dict_to_fields):
        'top_ad'    — wyróżnione na górze listy (płatne)
        'highlight' — podświetlone tło (płatne)
        'urgent'    — pilne (płatne)
        'premium'   — strona premium (płatne)
        'unknown'   — sygnał promocji bez rozpoznanego typu

    Primary strategy: OLX adds ?search_reason=search%7Cpromoted to promoted URLs.
    Fallback strategies: badges, CSS classes, text markers, data attributes.

    Returns: {
        'is_promoted': bool,
        'promotion_type': str | None,
        'confidence': float   # 0.0 – 1.0
    }
    """
    signals = []

    # STRATEGIA 0: URL parameter — najsilniejszy sygnał dla OLX
    for link in card.select('a[href*="/d/oferta/"]'):
        href = link.get('href', '')
        if 'search_reason=search%7Cpromoted' in href:
            signals.append(('url_promoted', 1.0))
            break

    # STRATEGIA 1: data-testid
    if card.select_one('[data-testid="adCard-featured"]'):
        signals.append(('testid_featured', 1.0))
    if card.select_one('[data-testid="listing-ad-badge"]'):
        signals.append(('testid_ad_badge', 0.9))

    # STRATEGIA 2: CSS classes
    element_classes = ' '.join(card.get('class', [])).lower()
    if any(kw in element_classes for kw in ['featured', 'promoted', 'top-ad', 'premium', 'vip']):
        signals.append(('css_top_ad', 0.8))
    if 'highlight' in element_classes or 'wyroznie' in element_classes:
        signals.append(('css_highlight', 0.8))

    # STRATEGIA 3: Text badges
    text_content = card.get_text()
    if any(b in text_content for b in ['Wyróżnione', 'Promowane', 'TOP']):
        signals.append(('text_top_ad', 0.85))
    if 'Pilne' in text_content:
        signals.append(('text_urgent', 0.85))
    if 'Premium' in text_content:
        signals.append(('text_premium', 0.85))

    # STRATEGIA 4: data attributes
    if card.get('data-promoted') or card.get('data-featured'):
        signals.append(('data_attribute', 1.0))

    if not signals:
        return {'is_promoted': False, 'promotion_type': None, 'confidence': 1.0}

    max_confidence = max(s[1] for s in signals)
    signal_types = [s[0] for s in signals]

    # Mapowanie sygnałów → canonical type (zgodny z promotion_dict_to_fields)
    if any(s in signal_types for s in ('url_promoted', 'testid_featured', 'testid_ad_badge',
                                        'css_top_ad', 'text_top_ad', 'data_attribute')):
        promo_type = 'top_ad'
    elif 'text_urgent' in signal_types:
        promo_type = 'urgent'
    elif 'text_premium' in signal_types:
        promo_type = 'premium'
    elif 'css_highlight' in signal_types:
        promo_type = 'highlight'
    else:
        promo_type = 'unknown'

    return {'is_promoted': True, 'promotion_type': promo_type, 'confidence': max_confidence}


def parse_card(card):
    """Parse a single OLX listing card. Works for both profile and category pages."""
    title = ""
    href = ""
    
    # Collect all links to offers
    offer_links = card.select('a[href*="/d/oferta/"]')
    
    # Try to find a link with meaningful text (title)
    for link in offer_links:
        txt = link.get_text(strip=True)
        current_href = link.get("href", "")
        
        # Save first href we find
        if current_href and not href:
            href = current_href
        
        # If this link has meaningful text, use it as title
        if txt and len(txt) > 3:
            title = txt
            href = current_href  # Update href to match title source
            break
    
    if not href:
        return None

    full_url = href if href.startswith("http") else f"https://www.olx.pl{href}"

    # If still no title, try to extract from URL
    if not title:
        url_match = re.search(r"/oferta/(.+?)-CID", href)
        if url_match:
            title = url_match.group(1).replace("-", " ").title()

    if not title:
        return None

    # Price via data-testid="ad-price"
    price_text = ""
    price_el = card.select_one('[data-testid="ad-price"]')
    if price_el:
        price_text = price_el.get_text(strip=True)

    # Date and location
    date_text = ""
    location_text = ""

    for el in card.find_all(["p", "span"]):
        txt = el.get_text(strip=True)
        if not txt or len(txt) > 120:
            continue
        txt_lower = txt.lower()

        if any(kw in txt_lower for kw in DATE_KEYWORDS):
            if " - " in txt:
                parts = txt.split(" - ", 1)
                location_text = parts[0].strip()
                date_text = parts[1].strip()
            elif not date_text:
                date_text = txt
        elif txt in ["Lublin", "Lublin, lubelskie"] and not location_text:
            location_text = txt

    img = card.select_one("img")
    image_url = img.get("src", "") if img else ""
    
    # NEW: Detect promoted status
    promo_status = detect_promoted_status(card)

    return {
        "title": title,
        "price_text": price_text,
        "price": parse_price(price_text),
        "date_text": date_text,
        "location": location_text,
        "url": full_url,
        "listing_id": extract_listing_id(full_url),
        "image_url": image_url,
        # NEW: Promoted fields
        "is_promoted": promo_status['is_promoted'],
        "promotion_type": promo_status['promotion_type'],
        "promotion_confidence": promo_status['confidence'],
    }


def parse_listings_from_soup(soup):
    """Parse all listings from page. Tries multiple card selectors."""
    listings = []

    # Strategy 1: data-cy="l-card" (category pages)
    cards = soup.select('[data-cy="l-card"]')

    # Strategy 2: profile page cards
    if not cards:
        cards = soup.select("div.css-19pezs8")

    # Strategy 3: any div containing an offer link + price
    if not cards:
        seen = set()
        for link in soup.select('a[href*="/d/oferta/"]'):
            href = link.get("href", "")
            if href in seen:
                continue
            seen.add(href)
            container = link
            for _ in range(6):
                p = container.parent
                if not p:
                    break
                if p.select_one('[data-testid="ad-price"]'):
                    container = p
                    break
                container = p
            if container != link:
                cards.append(container)

    for card in cards:
        parsed = parse_card(card)
        if parsed:
            listings.append(parsed)

    return listings


def get_total_count_from_header(soup):
    for el in soup.find_all(string=re.compile(r"Znaleźliśmy\s+\d+")):
        match = re.search(r"Znaleźliśmy\s+(\d+)\s+ogłosze", el)
        if match:
            return int(match.group(1))
    return None


def get_next_page_url(soup, current_url):
    # Strategy 1: Try standard pagination selectors (works for category pages)
    for selector in ['[data-testid="pagination-forward"]', '[data-cy="pagination-forward"]']:
        pag = soup.select_one(selector)
        if pag:
            href = pag.get("href", "")
            if href:
                return href if href.startswith("http") else f"https://www.olx.pl{href}"
    
    # Strategy 2: For user profiles - find next page link with SVG arrow
    # Extract current page number from URL
    current_page = 1
    page_match = re.search(r'page=(\d+)', current_url)
    if page_match:
        current_page = int(page_match.group(1))
    
    expected_next = current_page + 1
    
    # Find links with page= parameter
    links = soup.find_all('a', href=lambda x: x and 'page=' in x)
    
    # Look for link with SVG (arrow) that leads to next page
    for link in links:
        href = link.get('href', '')
        text = link.get_text(strip=True)
        has_svg = bool(link.find('svg'))
        
        # If has SVG and no numeric text (arrow button), check page number
        if has_svg and not text.isdigit():
            match = re.search(r'page=(\d+)', href)
            if match and int(match.group(1)) == expected_next:
                return href if href.startswith('http') else f'https://www.olx.pl{href}'
    
    # Fallback: look for any link to expected next page
    for link in links:
        href = link.get('href', '')
        if f'page={expected_next}' in href:
            return href if href.startswith('http') else f'https://www.olx.pl{href}'
    
    return None



# ─── JSON-based Scraping (for user profiles) ────────────────────────────────

def parse_prerendered_state(html):
    """
    Extract listings from OLX's __PRERENDERED_STATE__ JavaScript variable.
    OLX user profile pages now use client-side rendering with data embedded as JSON.
    """
    # Find the JSON data in __PRERENDERED_STATE__
    # Pattern: window.__PRERENDERED_STATE__ = "..."; followed by window.__ or </script>
    match = re.search(r'window\.__PRERENDERED_STATE__\s*=\s*"(.*?)";\s*(?:window\.|</script>)', html, re.DOTALL)
    if not match:
        return None, None, None
    
    try:
        # Decode the escaped JSON string
        # OLX encodes the JSON with unicode escapes, and UTF-8 characters 
        # get double-encoded, so we need to:
        # 1. Decode unicode escapes
        # 2. Fix UTF-8 that was incorrectly decoded as Latin-1
        import codecs
        encoded_str = match.group(1)
        decoded = codecs.decode(encoded_str, 'unicode_escape')
        # Fix UTF-8 encoding: the string was treated as Latin-1 but is actually UTF-8
        try:
            decoded = decoded.encode('latin-1').decode('utf-8')
        except (UnicodeDecodeError, UnicodeEncodeError):
            pass  # Already correct encoding, skip fix
        data = json.loads(decoded)
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        log.warning(f"Failed to parse __PRERENDERED_STATE__: {e}")
        return None, None, None
    
    # Extract listings from userListing.adsOffers
    user_listing = data.get('userListing', {})
    ads_offers = user_listing.get('adsOffers', {})
    
    ads = ads_offers.get('data', [])
    metadata = ads_offers.get('metadata', {})
    links = ads_offers.get('links', {})
    
    total_count = metadata.get('total_elements') or metadata.get('visible_total_count')
    next_page_url = links.get('next', {}).get('href') if isinstance(links.get('next'), dict) else None
    
    listings = []
    for ad in ads:
        # Extract price from params array
        # Price structure: param['value']['value'] contains the numeric price
        price = None
        price_text = ""
        for param in ad.get('params', []):
            if param.get('key') == 'price':
                value_obj = param.get('value', {})
                if isinstance(value_obj, dict):
                    price_text = value_obj.get('label', '')
                    # Try to get numeric value directly
                    price_val = value_obj.get('value')
                    if price_val is not None:
                        try:
                            price = int(price_val)
                        except (ValueError, TypeError):
                            price = parse_price(price_text)
                    else:
                        price = parse_price(price_text)
                break
        
        # Extract dates
        created_time = ad.get('created_time', '')
        last_refresh = ad.get('last_refresh_time', '')
        
        # Format dates
        published = None
        refreshed = None
        if created_time:
            try:
                dt = datetime.fromisoformat(created_time.replace('Z', '+00:00'))
                published = dt.strftime('%Y-%m-%d')
            except:
                pass
        if last_refresh:
            try:
                dt = datetime.fromisoformat(last_refresh.replace('Z', '+00:00'))
                refreshed = dt.strftime('%Y-%m-%d')
            except:
                pass
        
        # Build date_text for compatibility
        date_text = ""
        if refreshed:
            date_text = f"Odświeżono {refreshed}"
        elif published:
            date_text = published
        
        # Extract image
        photos = ad.get('photos', [])
        image_url = photos[0].get('link', '') if photos else ""
        
        # Extract location
        location = ad.get('location', {})
        location_text = location.get('city', {}).get('name', '')
        region = location.get('region', {}).get('name', '')
        if region and location_text:
            location_text = f"{location_text}, {region}"
        
        url = ad.get('url', '')
        if not url.startswith('http'):
            url = f"https://www.olx.pl{url}"
        
        # Ujednolicenie ID: zawsze używaj alfanumerycznego ID z URL (np. 19Qbyj),
        # żeby był zgodny z ID wyciąganym przez HTML/Playwright scraper kategorii.
        url_id = extract_listing_id(url)
        listing = {
            "title": ad.get('title', ''),
            "price": price,
            "price_text": price_text,
            "url": url,
            "listing_id": url_id,
            "date_text": date_text,
            "published": published,
            "refreshed": refreshed,
            "last_refresh_timestamp": last_refresh or None,  # pełny ISO timestamp z OLX, do detekcji zmian w ciągu dnia
            "location": location_text,
            "image_url": image_url,
            **promotion_dict_to_fields(ad.get("promotion", {})),
        }
        
        if listing["title"]:  # Only add if has title
            listings.append(listing)
    
    return listings, total_count, next_page_url


def scrape_user_profile_json(profile_key, profile_config, session):
    """
    Scrape user profile by parsing __PRERENDERED_STATE__ JSON data.
    This is the new method for OLX user profiles which use client-side rendering.
    Falls back to API pagination for additional pages.
    """
    url = profile_config["url"]
    all_listings = []
    header_count = None
    page = 1
    max_pages = 20
    
    # Random initial delay
    time.sleep(random.uniform(1.5, 3.0))
    
    log.info(f"  [{profile_key}] Page {page}: {url}")
    
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as e:
        log.error(f"  [{profile_key}] HTTP error: {e}")
        return {
            "listings": [],
            "count": 0,
            "header_count": None,
            "pages_scraped": 0,
        }
    
    # Parse first page from __PRERENDERED_STATE__
    listings, total_count, next_api_url = parse_prerendered_state(resp.text)
    
    if listings is None:
        log.warning(f"  [{profile_key}] Could not parse __PRERENDERED_STATE__, trying DOM parsing")
        # Fallback to DOM parsing
        soup = BeautifulSoup(resp.text, "lxml")
        listings = parse_listings_from_soup(soup)
        total_count = get_total_count_from_header(soup)
    
    header_count = total_count
    log.info(f"  [{profile_key}] Page 1: {len(listings)} listings (total: {total_count})")
    all_listings.extend(listings)
    
    # Fetch additional pages via API
    while next_api_url and page < max_pages and len(all_listings) < (total_count or 999):
        page += 1
        time.sleep(random.uniform(1.5, 3.0))
        
        log.info(f"  [{profile_key}] API Page {page}: {next_api_url[:80]}...")
        
        try:
            api_resp = session.get(next_api_url, timeout=30)
            api_resp.raise_for_status()
            api_data = api_resp.json()
        except Exception as e:
            log.warning(f"  [{profile_key}] API error on page {page}: {e}")
            break
        
        ads = api_data.get('data', [])
        if not ads:
            break
        
        for ad in ads:
            # Extract price - same structure as in parse_prerendered_state
            price = None
            price_text = ""
            for param in ad.get('params', []):
                if param.get('key') == 'price':
                    value_obj = param.get('value', {})
                    if isinstance(value_obj, dict):
                        price_text = value_obj.get('label', '')
                        price_val = value_obj.get('value')
                        if price_val is not None:
                            try:
                                price = int(price_val)
                            except (ValueError, TypeError):
                                price = parse_price(price_text)
                        else:
                            price = parse_price(price_text)
                    break
            
            # Dates
            created_time = ad.get('created_time', '')
            last_refresh = ad.get('last_refresh_time', '')
            published = None
            refreshed = None
            if created_time:
                try:
                    dt = datetime.fromisoformat(created_time.replace('Z', '+00:00'))
                    published = dt.strftime('%Y-%m-%d')
                except:
                    pass
            if last_refresh:
                try:
                    dt = datetime.fromisoformat(last_refresh.replace('Z', '+00:00'))
                    refreshed = dt.strftime('%Y-%m-%d')
                except:
                    pass
            
            date_text = f"Odświeżono {refreshed}" if refreshed else (published or "")
            
            photos = ad.get('photos', [])
            image_url = photos[0].get('link', '') if photos else ""
            
            location = ad.get('location', {})
            location_text = location.get('city', {}).get('name', '')
            
            url = ad.get('url', '')
            if not url.startswith('http'):
                url = f"https://www.olx.pl{url}"
            
            listing = {
                "title": ad.get('title', ''),
                "price": price,
                "price_text": price_text,
                "url": url,
                "listing_id": extract_listing_id(url),
                "date_text": date_text,
                "published": published,
                "refreshed": refreshed,
                "last_refresh_timestamp": last_refresh or None,
                "location": location_text,
                "image_url": image_url,
                **promotion_dict_to_fields(ad.get("promotion", {})),
            }
            
            if listing["title"]:
                all_listings.append(listing)
        
        log.info(f"  [{profile_key}] Page {page}: {len(ads)} listings")
        
        # Get next page URL
        links = api_data.get('links', {})
        next_api_url = links.get('next', {}).get('href') if isinstance(links.get('next'), dict) else None
    
    # Deduplicate
    seen_ids = set()
    unique = []
    for listing in all_listings:
        lid = listing["listing_id"]
        if lid not in seen_ids:
            seen_ids.add(lid)
            unique.append(listing)
    
    return {
        "listings": unique,
        "count": len(unique),
        "header_count": header_count,
        "pages_scraped": page,
    }


# ─── OLX API Scraping (user profiles) ────────────────────────────────────────

def scrape_user_via_api(profile_key, profile_config):
    """
    Scrape a user profile via OLX REST API (api/v1/offers?user_id=UUID).
    Works reliably from GitHub Actions IPs — no browser needed.
    Requires 'uuid' key in profile_config.
    """
    uuid = profile_config.get("uuid")
    if not uuid:
        log.error(f"  [{profile_key}] No UUID configured, cannot use API scraping")
        return {"listings": [], "count": 0, "header_count": None, "pages_scraped": 0}

    s = get_api_session()
    all_ads = []
    offset = 0
    limit = 10
    total = None
    page_num = 1
    max_pages = 50

    while page_num <= max_pages:
        url = (
            f"https://www.olx.pl/api/v1/offers"
            f"?offset={offset}&limit={limit}&category_id=0"
            f"&sort_by=created_at%3Adesc&user_id={uuid}"
        )
        log.info(f"  [{profile_key}] API page {page_num} (offset={offset})")
        try:
            r = s.get(url, timeout=20)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            log.error(f"  [{profile_key}] API error on page {page_num}: {e}")
            break

        ads = data.get("data", [])
        meta = data.get("metadata", {})
        links = data.get("links", {})

        if total is None:
            total = meta.get("total_elements", 0)
            log.info(f"  [{profile_key}] Total (API): {total}")

        log.info(f"  [{profile_key}] Page {page_num}: {len(ads)} listings")
        all_ads.extend(ads)

        # Stop if we have everything or no next page
        next_link = links.get("next")
        has_next = bool(next_link and isinstance(next_link, dict) and next_link.get("href"))
        if not has_next or len(all_ads) >= (total or 0):
            break

        offset += limit
        page_num += 1
        time.sleep(random.uniform(0.8, 1.5))

    listings = _parse_ads_json(all_ads)

    # Deduplicate
    seen_ids = set()
    unique = []
    for listing in listings:
        lid = listing["listing_id"]
        if lid not in seen_ids:
            seen_ids.add(lid)
            unique.append(listing)

    return {
        "listings": unique,
        "count": len(unique),
        "header_count": total,
        "pages_scraped": page_num,
    }


# ─── Playwright Scraping (kategoria) ─────────────────────────────────────────

def _parse_ads_json(ads):
    """Convert OLX adsOffers data[] items to our listing dict format."""
    listings = []
    for ad in ads:
        price = None
        price_text = ""
        for param in ad.get("params", []):
            if param.get("key") == "price":
                val = param.get("value", {})
                if isinstance(val, dict):
                    price_text = val.get("label", "")
                    pv = val.get("value")
                    if pv is not None:
                        try:
                            price = int(pv)
                        except (ValueError, TypeError):
                            price = parse_price(price_text)
                    else:
                        price = parse_price(price_text)
                break

        created_time = ad.get("created_time", "")
        last_refresh = ad.get("last_refresh_time", "")
        published = None
        refreshed = None
        if created_time:
            try:
                published = datetime.fromisoformat(created_time.replace("Z", "+00:00")).strftime("%Y-%m-%d")
            except Exception:
                pass
        if last_refresh:
            try:
                refreshed = datetime.fromisoformat(last_refresh.replace("Z", "+00:00")).strftime("%Y-%m-%d")
            except Exception:
                pass

        date_text = f"Odświeżono {refreshed}" if refreshed else (published or "")

        photos = ad.get("photos", [])
        image_url = photos[0].get("link", "") if photos else ""

        location = ad.get("location", {})
        location_text = location.get("city", {}).get("name", "")

        url = ad.get("url", "")
        if not url.startswith("http"):
            url = f"https://www.olx.pl{url}"

        if ad.get("title"):
            listings.append({
                "title": ad["title"],
                "price": price,
                "price_text": price_text,
                "url": url,
                "listing_id": extract_listing_id(url),
                "date_text": date_text,
                "published": published,
                "refreshed": refreshed,
                "last_refresh_timestamp": last_refresh or None,
                "location": location_text,
                "image_url": image_url,
                **promotion_dict_to_fields(ad.get("promotion", {})),
            })
    return listings


def _scrape_one_profile_playwright(page_obj, profile_key, profile_config):
    """
    Scrape a single profile using an already-open Playwright page object.

    - User profiles: load page → extract __PRERENDERED_STATE__ via JS evaluate
      (JS engine already parsed/executed it — no regex needed).
      Pagination by incrementing ?page=N in URL while next link exists.
    - Category pages: wait for [data-cy="l-card"] → parse DOM → follow
      pagination-forward link.
    """
    url = profile_config["url"]
    is_category = profile_config.get("is_category", False)
    all_listings = []
    header_count = None
    page_num = 1
    max_pages = 50

    while url and page_num <= max_pages:
        log.info(f"  [{profile_key}] Page {page_num}: {url}")
        try:
            page_obj.goto(url, wait_until="domcontentloaded", timeout=45000)
            time.sleep(random.uniform(2.0, 3.5))
        except PlaywrightTimeout:
            log.warning(f"  [{profile_key}] Timeout on page {page_num}, stopping")
            break
        except Exception as e:
            log.error(f"  [{profile_key}] Navigation error page {page_num}: {e}")
            break

        if not is_category:
            # ── User profile: wait for __PRERENDERED_STATE__ then extract ──
            try:
                # Poll until window.__PRERENDERED_STATE__ is populated by JS
                page_obj.wait_for_function(
                    "() => window.__PRERENDERED_STATE__ !== undefined && window.__PRERENDERED_STATE__ !== null",
                    timeout=20000,
                )
            except PlaywrightTimeout:
                log.warning(f"  [{profile_key}] __PRERENDERED_STATE__ never populated (timeout)")
                break
            except Exception as e:
                log.warning(f"  [{profile_key}] wait_for_function error: {e}")
                break

            try:
                raw = page_obj.evaluate("() => JSON.stringify(window.__PRERENDERED_STATE__)")
            except Exception as e:
                log.warning(f"  [{profile_key}] JS eval error: {e}")
                break

            if not raw or raw in ("null", "undefined", "None"):
                log.warning(f"  [{profile_key}] __PRERENDERED_STATE__ is empty")
                break

            try:
                data = json.loads(raw)
                if isinstance(data, str):
                    data = json.loads(data)
            except json.JSONDecodeError as e:
                log.warning(f"  [{profile_key}] JSON decode error: {e}")
                break

            ads_offers = data.get("userListing", {}).get("adsOffers", {})
            ads = ads_offers.get("data", [])
            metadata = ads_offers.get("metadata", {})
            links = ads_offers.get("links", {})

            if page_num == 1:
                header_count = metadata.get("total_elements") or metadata.get("visible_total_count")
                log.info(f"  [{profile_key}] Total (header): {header_count}")

            page_listings = _parse_ads_json(ads)
            log.info(f"  [{profile_key}] Page {page_num}: {len(page_listings)} listings")
            all_listings.extend(page_listings)

            next_link = links.get("next")
            has_next = bool(next_link and (isinstance(next_link, dict) and next_link.get("href") or isinstance(next_link, str)))

            if has_next:
                # Build next HTML page URL by incrementing page= param
                current_page_match = re.search(r"[?&]page=(\d+)", url)
                current_page_n = int(current_page_match.group(1)) if current_page_match else 1
                next_page_n = current_page_n + 1
                if "page=" in url:
                    next_url = re.sub(r"page=\d+", f"page={next_page_n}", url)
                elif "?" in url:
                    next_url = url + f"&page={next_page_n}"
                else:
                    next_url = url + f"?page={next_page_n}"
                url = next_url
                page_num += 1
                time.sleep(random.uniform(1.5, 3.0))
            else:
                break

        else:
            # ── Category page: parse DOM ──────────────────────────────────
            try:
                page_obj.wait_for_selector("[data-cy='l-card']", timeout=15000)
            except PlaywrightTimeout:
                log.warning(f"  [{profile_key}] No cards found on page {page_num}")
                break

            html = page_obj.content()
            soup = BeautifulSoup(html, "lxml")

            if page_num == 1:
                header_count = get_total_count_from_header(soup)
                log.info(f"  [{profile_key}] Header count: {header_count}")

            page_listings = parse_listings_from_soup(soup)
            log.info(f"  [{profile_key}] Page {page_num}: {len(page_listings)} listings")

            if not page_listings:
                break

            all_listings.extend(page_listings)

            next_url = get_next_page_url(soup, url)
            if next_url:
                url = next_url
                page_num += 1
                time.sleep(random.uniform(2.0, 4.0))
            else:
                break

    seen_ids = set()
    unique = []
    for listing in all_listings:
        lid = listing["listing_id"]
        if lid not in seen_ids:
            seen_ids.add(lid)
            unique.append(listing)

    return {
        "listings": unique,
        "count": len(unique),
        "header_count": header_count,
        "pages_scraped": page_num,
    }


def scrape_with_playwright_all(profiles):
    """
    Scrape all profiles using the best method per profile type:
    - User profiles (is_category=False): OLX REST API via requests — works from any IP.
    - Category pages (is_category=True): Playwright headless Chromium — handles SSR DOM.
    Returns {profile_key: result_dict}.
    """
    results = {}

    # Split profiles by type
    category_profiles = {pk: cfg for pk, cfg in profiles.items() if cfg.get("is_category")}
    user_profiles = {pk: cfg for pk, cfg in profiles.items() if not cfg.get("is_category")}

    # ── User profiles via API (requests, no browser needed) ───────────────────
    for pk, cfg in user_profiles.items():
        log.info(f"[SCAN] Profile: {pk} (API)")
        profile_start = time.time()
        try:
            result = scrape_user_via_api(pk, cfg)

            scraped = result["count"]
            header = result["header_count"]
            tolerance = 2
            header_match = header is None or abs(scraped - header) <= tolerance

            if header_match:
                result["crosscheck"] = "passed"
                result["crosscheck_detail"] = f"scraped={scraped}, header={header}"
                log.info(f"[CROSSCHECK] {pk}: PASS (scraped={scraped}, header={header})")
            else:
                log.info(f"[CROSSCHECK] {pk}: MISMATCH scraped={scraped} vs header={header}, retrying...")
                time.sleep(random.uniform(3, 5))
                result2 = scrape_user_via_api(pk, cfg)
                c2 = result2["count"]
                if c2 > scraped:
                    result2["crosscheck"] = "passed_retry"
                    result2["crosscheck_detail"] = f"1st={scraped}, 2nd={c2}, header={header}"
                    result = result2
                else:
                    result["crosscheck"] = "best_of_two"
                    result["crosscheck_detail"] = f"1st={scraped}, 2nd={c2}, header={header}"

            result["duration_seconds"] = round(time.time() - profile_start, 1)
            results[pk] = result
            log.info(f"[OK] {pk}: {result['count']} listings ({result['crosscheck']}) [{result['duration_seconds']}s]")
        except Exception as e:
            log.error(f"[ERROR] {pk}: {e}")
            results[pk] = {
                "listings": [], "count": 0, "header_count": None,
                "crosscheck": "error", "crosscheck_detail": str(e), "pages_scraped": 0,
                "duration_seconds": round(time.time() - profile_start, 1),
            }
        time.sleep(random.uniform(1, 2))

    # ── Category pages via Playwright ─────────────────────────────────────────
    if category_profiles:
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"],
                )
                context = browser.new_context(
                    user_agent=random.choice(USER_AGENTS),
                    viewport={"width": 1920, "height": 1080},
                    locale="pl-PL",
                    timezone_id="Europe/Warsaw",
                )
                page_obj = context.new_page()

                for pk, cfg in category_profiles.items():
                    log.info(f"[SCAN] Profile: {pk} (Playwright)")
                    profile_start = time.time()
                    try:
                        result = _scrape_one_profile_playwright(page_obj, pk, cfg)

                        scraped = result["count"]
                        header = result["header_count"]
                        tolerance = 10
                        header_match = header is None or abs(scraped - header) <= tolerance

                        if header_match:
                            result["crosscheck"] = "passed"
                            result["crosscheck_detail"] = f"scraped={scraped}, header={header}"
                            log.info(f"[CROSSCHECK] {pk}: PASS (scraped={scraped}, header={header})")
                        else:
                            log.info(f"[CROSSCHECK] {pk}: MISMATCH scraped={scraped} vs header={header}, retrying...")
                            time.sleep(random.uniform(4, 7))
                            result2 = _scrape_one_profile_playwright(page_obj, pk, cfg)
                            c2 = result2["count"]
                            if c2 > scraped:
                                result2["crosscheck"] = "passed_retry"
                                result2["crosscheck_detail"] = f"1st={scraped}, 2nd={c2}, header={header}"
                                result = result2
                            else:
                                result["crosscheck"] = "best_of_two"
                                result["crosscheck_detail"] = f"1st={scraped}, 2nd={c2}, header={header}"

                        result["duration_seconds"] = round(time.time() - profile_start, 1)
                        results[pk] = result
                        log.info(f"[OK] {pk}: {result['count']} listings ({result['crosscheck']}) [{result['duration_seconds']}s]")
                    except Exception as e:
                        log.error(f"[ERROR] {pk}: {e}")
                        results[pk] = {
                            "listings": [], "count": 0, "header_count": None,
                            "crosscheck": "error", "crosscheck_detail": str(e), "pages_scraped": 0,
                            "duration_seconds": round(time.time() - profile_start, 1),
                        }

                browser.close()

        except Exception as e:
            log.error(f"[PLAYWRIGHT] Fatal browser error: {e}")
            for pk in category_profiles:
                if pk not in results:
                    results[pk] = {
                        "listings": [], "count": 0, "header_count": None,
                        "crosscheck": "error", "crosscheck_detail": str(e), "pages_scraped": 0,
                        "duration_seconds": 0,
                    }

    return results


def scrape_with_crosscheck(profile_key, profile_config):
    """Single-profile shim for backward compatibility."""
    results = scrape_with_playwright_all({profile_key: profile_config})
    return results[profile_key]


# ─── Excel Operations ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)
UP_FONT = Font(name="Arial", size=10, color="00B050")
DOWN_FONT = Font(name="Arial", size=10, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER


def style_data_cell(cell, font=None):
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def get_or_create_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
        # Update headers if they changed (e.g. new columns added)
        existing_headers = [ws.cell(row=1, column=ci).value for ci in range(1, len(headers) + 1)]
        if existing_headers != headers:
            for ci, h in enumerate(headers, 1):
                ws.cell(row=1, column=ci, value=h)
            style_header_row(ws, 1, len(headers))
        return ws
    ws = wb.create_sheet(name)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header_row(ws, 1, len(headers))
    return ws


# ─── Excel generowany na żądanie (z JSON + ledger) ──────────────────────────

def _load_daily_ledger(ledger_path=HISTORY_LEDGER):
    """Wczytuje append-only ledger trendu dziennego (NDJSON). Pusta lista jeśli brak."""
    rows = []
    if os.path.exists(ledger_path):
        with open(ledger_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    return rows


def build_excel_from_data(output_path, json_path=JSON_PATH, ledger_path=HISTORY_LEDGER):
    """
    Buduje kompletny arkusz Excela z `dashboard_data.json` (pełne dane per-ogłoszenie:
    current + archived + price_history) oraz z ledgera trendu dziennego (NDJSON).
    Nie modyfikuje żadnych plików źródłowych — zapisuje TYLKO do `output_path`.
    Zwraca `output_path`. Używane do załącznika maila (nie commitowane do repo).
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    profiles = data.get("profiles", {})
    ledger = _load_daily_ledger(ledger_path)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Arkusze per profil: aktywne + archiwalne ogłoszenia ──
    prof_headers = [
        "Status", "Tytuł", "Cena (zł)", "Pierwsza cena", "Zmiana ceny",
        "Promowane", "Typ prom.", "Dni prom.", "Liczba odświeżeń",
        "Data publikacji", "Data odświeżenia", "Pierwszy raz", "Ostatni raz",
        "Data archiwizacji", "URL", "ID ogłoszenia",
    ]
    for pk, pd_ in profiles.items():
        ws = get_or_create_sheet(wb, pk[:31], prof_headers)
        rows = ([("aktywne", l) for l in pd_.get("current_listings", [])] +
                [("archiwum", l) for l in pd_.get("archived_listings", [])])
        r = 2
        for status, l in rows:
            ws.cell(row=r, column=1, value=status)
            ws.cell(row=r, column=2, value=l.get("title"))
            ws.cell(row=r, column=3, value=l.get("price"))
            ws.cell(row=r, column=4, value=l.get("first_price"))
            pc = l.get("price_change")
            ch_cell = ws.cell(row=r, column=5, value=pc)
            ws.cell(row=r, column=6, value="✓" if l.get("is_promoted") else "—")
            ws.cell(row=r, column=7, value=l.get("promotion_type") or "—")
            ws.cell(row=r, column=8, value=l.get("promoted_days_current", 0))
            ws.cell(row=r, column=9, value=l.get("refresh_count", 0))
            ws.cell(row=r, column=10, value=l.get("published") or "")
            ws.cell(row=r, column=11, value=l.get("refreshed") or "")
            ws.cell(row=r, column=12, value=l.get("first_seen") or "")
            ws.cell(row=r, column=13, value=l.get("last_seen") or "")
            ws.cell(row=r, column=14, value=l.get("archived_date") or "")
            ws.cell(row=r, column=15, value=l.get("url"))
            ws.cell(row=r, column=16, value=l.get("id"))
            for c in range(1, 17):
                f = (UP_FONT if (c == 5 and isinstance(pc, (int, float)) and pc > 0)
                     else DOWN_FONT if (c == 5 and isinstance(pc, (int, float)) and pc < 0)
                     else DATA_FONT)
                style_data_cell(ws.cell(row=r, column=c), f)
            r += 1
        for idx, w in enumerate([10, 50, 12, 12, 12, 11, 14, 9, 12, 16, 16, 16, 16, 16, 60, 15], 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    # ── historia_cen: wszystkie zdarzenia zmiany ceny ──
    ph = ["Data", "Profil", "ID ogłoszenia", "Tytuł", "Stara cena", "Nowa cena", "Zmiana", "URL"]
    ws_p = get_or_create_sheet(wb, "historia_cen", ph)
    r = 2
    price_rows = []
    for pk, pd_ in profiles.items():
        # mapa id -> (title, url) z current + archived
        meta = {}
        for l in pd_.get("current_listings", []) + pd_.get("archived_listings", []):
            meta[l.get("id")] = (l.get("title"), l.get("url"))
        for lid, events in pd_.get("price_history", {}).items():
            title, url = meta.get(lid, (None, None))
            for ev in events:
                price_rows.append((ev.get("date"), pk, lid, title,
                                   ev.get("old_price"), ev.get("new_price"),
                                   ev.get("change"), url))
    price_rows.sort(key=lambda x: (x[0] or "", x[1], str(x[2])))
    for date, pk, lid, title, op, np_, chg, url in price_rows:
        ws_p.cell(row=r, column=1, value=date)
        ws_p.cell(row=r, column=2, value=pk)
        ws_p.cell(row=r, column=3, value=lid)
        ws_p.cell(row=r, column=4, value=title)
        ws_p.cell(row=r, column=5, value=op)
        ws_p.cell(row=r, column=6, value=np_)
        ws_p.cell(row=r, column=7, value=chg)
        ws_p.cell(row=r, column=8, value=url)
        for c in range(1, 9):
            f = (DOWN_FONT if (c == 7 and isinstance(chg, (int, float)) and chg < 0)
                 else UP_FONT if (c == 7 and isinstance(chg, (int, float)) and chg > 0)
                 else DATA_FONT)
            style_data_cell(ws_p.cell(row=r, column=c), f)
        r += 1
    for idx, w in enumerate([12, 18, 15, 50, 12, 12, 12, 60], 1):
        ws_p.column_dimensions[get_column_letter(idx)].width = w

    # ── trend_dzienny: WIECZNY trend z ledgera (1 wiersz = 1 skan) ──
    th = ["Data", "Godzina", "Profil", "Liczba ogłoszeń", "Crosscheck", "Zmiana vs poprzedni"]
    ws_t = get_or_create_sheet(wb, "trend_dzienny", th)
    r = 2
    for row in ledger:
        ws_t.cell(row=r, column=1, value=row.get("date"))
        ws_t.cell(row=r, column=2, value=row.get("time"))
        ws_t.cell(row=r, column=3, value=row.get("profile"))
        ws_t.cell(row=r, column=4, value=row.get("count"))
        ws_t.cell(row=r, column=5, value=row.get("crosscheck"))
        chg = row.get("change")
        ws_t.cell(row=r, column=6, value=chg)
        for c in range(1, 7):
            f = (UP_FONT if (c == 6 and isinstance(chg, (int, float)) and chg > 0)
                 else DOWN_FONT if (c == 6 and isinstance(chg, (int, float)) and chg < 0)
                 else DATA_FONT)
            style_data_cell(ws_t.cell(row=r, column=c), f)
        r += 1
    for idx, w in enumerate([12, 10, 20, 16, 14, 18], 1):
        ws_t.column_dimensions[get_column_letter(idx)].width = w

    # ── podsumowanie: aktualny stan per profil ──
    sh = ["Profil", "Label", "Aktywne", "Archiwalne", "Ostatni skan"]
    ws_s = wb.create_sheet("podsumowanie")
    for ci, h in enumerate(sh, 1):
        ws_s.cell(row=1, column=ci, value=h)
    style_header_row(ws_s, 1, len(sh))
    ri = 2
    last_scan = data.get("last_scan", "")
    for pk, pd_ in profiles.items():
        ws_s.cell(row=ri, column=1, value=pk)
        ws_s.cell(row=ri, column=2, value=pd_.get("label", PROFILES.get(pk, {}).get("label", pk)))
        ws_s.cell(row=ri, column=3, value=len(pd_.get("current_listings", [])))
        ws_s.cell(row=ri, column=4, value=len(pd_.get("archived_listings", [])))
        ws_s.cell(row=ri, column=5, value=last_scan)
        for c in range(1, 6):
            style_data_cell(ws_s.cell(row=ri, column=c))
        ri += 1
    for idx, w in enumerate([20, 30, 12, 12, 20], 1):
        ws_s.column_dimensions[get_column_letter(idx)].width = w

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    log.info(f"Excel (z JSON+ledger) zapisany: {output_path}")
    return output_path


def append_history(scan_results, scan_timestamp, ledger_path=HISTORY_LEDGER, json_path=JSON_PATH):
    """
    Dopisuje (APPEND-ONLY) do ledgera trendu dziennego po jednej linii na profil.
    Nigdy nie nadpisuje istniejących linii. Zachowuje tę samą ochronę co
    `generate_dashboard_json`: przy błędzie scrapera (a mając wcześniejsze dane)
    NIE dopisuje fałszywego zera — pomija profil, żeby nie zafałszować trendu.

    WAŻNE: wywoływane PO `generate_dashboard_json`, więc `dashboard_data.json`
    jest już zaktualizowany (chroni current_listings przy błędach).
    """
    os.makedirs(os.path.dirname(ledger_path), exist_ok=True)

    # Ostatni znany count per profil (do pola `change`)
    last_count = {}
    for row in _load_daily_ledger(ledger_path):
        last_count[row["profile"]] = row["count"]

    # Liczba aktualnych ogłoszeń per profil z JSON (już zaktualizowany)
    prior_listing_count = {}
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                d = json.load(f)
            for pk, pd_ in d.get("profiles", {}).items():
                prior_listing_count[pk] = len(pd_.get("current_listings", []))
        except (json.JSONDecodeError, IOError):
            pass

    date = scan_timestamp.strftime("%Y-%m-%d")
    time_s = scan_timestamp.strftime("%H:%M")
    appended = 0
    with open(ledger_path, "a", encoding="utf-8") as f:
        for pk, result in scan_results.items():
            crosscheck = result.get("crosscheck", "")
            header_count = result.get("header_count")
            count = result.get("count", 0)
            is_scraper_error = (crosscheck == "error"
                                or (count == 0 and header_count is None)
                                or is_header_shortfall(result))
            if is_scraper_error and prior_listing_count.get(pk, 0) > 0:
                log.warning(f"[{pk}] Pomijam wpis do ledgera — błąd scrapera "
                            f"(crosscheck={crosscheck}, header={header_count})")
                continue
            prev = last_count.get(pk)
            change = (count - prev) if prev is not None else 0
            rec = {"date": date, "time": time_s, "profile": pk, "count": count,
                   "crosscheck": crosscheck, "change": change, "source": "scan"}
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            last_count[pk] = count
            appended += 1
    log.info(f"Ledger: dopisano {appended} wpisów -> {ledger_path}")
    return appended


def generate_trend_full(ledger_path=HISTORY_LEDGER, out_path=None):
    """
    Buduje lekki `docs/api/trend_full.json` z ledgera — PEŁNA historia liczby ogłoszeń
    (1 punkt/dzień/profil = ostatni skan danego dnia). Dashboard pobiera go tylko gdy
    użytkownik włączy widok 'cała historia' na wykresie trendu. Pozostałe metryki
    (mediana itd.) nie są dostępne historycznie — żyją w daily_counts (90 dni).
    """
    out_path = out_path or os.path.join(API_DIR, "trend_full.json")
    ledger = _load_daily_ledger(ledger_path)

    # profile -> {date: (time, count)} ; przy wielu skanach w dniu bierzemy ostatni (max time)
    by_prof = {}
    for r in ledger:
        p, d, t, c = r["profile"], r["date"], r.get("time", ""), r["count"]
        day_map = by_prof.setdefault(p, {})
        prev = day_map.get(d)
        if prev is None or t >= prev[0]:
            day_map[d] = (t, c)

    profiles = {p: [{"date": d, "count": dm[d][1]} for d in sorted(dm)]
                for p, dm in by_prof.items()}

    # Odpływ ofert („ile znika z rynku") — pełna historia per profil.
    # Dla każdego dnia liczymy ogłoszenia zarchiwizowane tego dnia (data-part
    # `archived_date`). Źródłem jest `archived_listings` w dashboard_data.json,
    # które NIE ma limitu 90 dni (w przeciwieństwie do `daily_counts.removed`),
    # więc daje pełną historię. Konwencja liczenia po dniu archiwizacji jest
    # spójna z metryką „Zniknęło" (`daily_counts.removed`) na dashboardzie.
    outflow = {}
    try:
        data = load_existing_json()
        for pk, pdata in (data.get("profiles") or {}).items():
            day_counts = {}
            for l in pdata.get("archived_listings", []):
                ad = l.get("archived_date")
                if not ad:
                    continue
                day = ad[:10]  # 'YYYY-MM-DD'
                if len(day) == 10:
                    day_counts[day] = day_counts.get(day, 0) + 1
            if day_counts:
                outflow[pk] = [{"date": d, "count": day_counts[d]} for d in sorted(day_counts)]
    except Exception as e:
        log.warning(f"Nie udało się policzyć odpływu ofert do trend_full.json: {e}")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    payload = {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "profiles": profiles,
        "outflow": outflow,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, sort_keys=True)
    total = sum(len(v) for v in profiles.values())
    out_total = sum(len(v) for v in outflow.values())
    log.info(f"trend_full.json zapisany: {total} punktów count + {out_total} punktów odpływu "
             f"({len(profiles)} profili) -> {out_path}")
    return out_path


# ─── Filtrowanie cenowych odstających wartości ──────────────────────────────

def filter_price_outliers(scan_results, multiplier=PRICE_OUTLIER_MULTIPLIER):
    """
    Odrzuca ze skanu ogłoszenia, których cena >= `multiplier` x średnia cena
    POZOSTAŁYCH ogłoszeń w danym profilu (typowo literówka w cenie albo
    ogłoszenie, które nie jest pokojem). Modyfikuje `result["listings"]`
    i `result["count"]` w miejscu, więc dane nigdy nie trafiają do
    dashboard_data.json / ledgera / API.
    Wymaga min. 3 wycenionych ogłoszeń w profilu, żeby średnia miała sens.

    Średnia liczona jest metodą "leave-one-out" (bez ceny sprawdzanego
    ogłoszenia) — inaczej sam outlier zawyżałby własną średnią odniesienia
    i nigdy nie przekroczyłby progu (np. przy 4 cenach: 1000, 1200, 900,
    126065 zł, średnia ze WSZYSTKICH to >32000 zł, więc 126065 zł nie
    przekroczyłoby 10x tej średniej; średnia z pozostałych trzech to
    ~1033 zł, więc 126065 zł jednoznacznie przekracza 10x próg).
    Filtrowanie jest iteracyjne — po odrzuceniu outlierów w danej rundzie
    średnia dla pozostałych jest przeliczana i sprawdzana ponownie.
    """
    for pk, result in scan_results.items():
        listings = result.get("listings", [])
        removed = []
        remaining = list(listings)

        while True:
            priced_items = [(l, l["price"]) for l in remaining if l.get("price") and l["price"] > 0]
            if len(priced_items) < 3:
                break
            total = sum(p for _, p in priced_items)
            n = len(priced_items)

            round_removed = []
            round_removed_ids = set()
            for l, p in priced_items:
                others_avg = (total - p) / (n - 1)
                if p >= others_avg * multiplier:
                    round_removed.append((l, others_avg))
                    round_removed_ids.add(id(l))

            if not round_removed:
                break
            removed.extend(round_removed)
            remaining = [l for l in remaining if id(l) not in round_removed_ids]

        if removed:
            for l, others_avg in removed:
                log.warning(
                    f"[{pk}] Pominięto outlier cenowy (>= {multiplier}x średnia "
                    f"{others_avg:.0f} zł): '{l.get('title', '')[:60]}' "
                    f"{l.get('price_text', '')} — {l.get('url', '')}"
                )
            result["listings"] = remaining
            result["count"] = len(remaining)

    return scan_results


# ─── JSON for Dashboard ─────────────────────────────────────────────────────

def load_existing_json():
    if os.path.exists(JSON_PATH):
        try:
            with open(JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"profiles": {}, "scan_history": [], "last_scan": None}


def recompute_daily_refresh_reactivation(pd_):
    """Przelicza `refreshed_count` i `reactivated_count` we WSZYSTKICH wpisach
    `daily_counts` profilu jako projekcję historii ogłoszeń
    (`refresh_history`/`reactivation_history`, current + archived).

    Powód: skan leci raz dziennie rano, więc odświeżenie po godzinie skanu jest
    wykrywane dopiero następnego dnia — z wczorajszą datą `refreshed_at`. Stare
    liczenie ("wpisy z refreshed_at == dzisiaj") nigdy nie zliczało takich
    eventów (w danych ~54% eventów!). Projekcja przypisuje event do właściwego
    dnia wstecz (backfill) i jest idempotentna przy wielu skanach dziennie
    (eventy w historii są deduplikowane per data).

    Event z datą bez wpisu w daily_counts (dzień bez skanu) doliczany jest do
    najbliższego późniejszego wpisu, żeby nie przepadł; eventy sprzed okna
    90 dni są poza wykresem i pomijane.

    WAŻNE: liczniki są tylko PODNOSZONE (max ze starej i przeliczonej wartości),
    nigdy zmniejszane. Historie ogłoszeń bywają tracone przez znany bug cichego
    gubienia ogłoszeń przy rotacji wyników OLX (patrz CLAUDE.md §7) — czysta
    projekcja wyzerowałaby wtedy prawdziwe, wcześniej policzone dni.
    """
    dc = pd_.get("daily_counts", [])
    if not dc:
        return
    dates = sorted(e["date"] for e in dc if e.get("date"))
    if not dates:
        return
    date_set = set(dates)

    def bucket_for(day):
        if not day or day < dates[0]:
            return None  # sprzed okna daily_counts — poza wykresem
        if day in date_set:
            return day
        return next((d for d in dates if d > day), dates[-1])

    refresh_by_day = defaultdict(int)
    react_by_day = defaultdict(int)
    for l in pd_.get("current_listings", []) + pd_.get("archived_listings", []):
        for h in l.get("refresh_history", []):
            b = bucket_for(str(h.get("refreshed_at", ""))[:10])
            if b:
                refresh_by_day[b] += 1
        for h in l.get("reactivation_history", []):
            b = bucket_for(str(h.get("reactivated_at", ""))[:10])
            if b:
                react_by_day[b] += 1

    for e in dc:
        d = e.get("date")
        if not d:
            continue
        e["refreshed_count"] = max(int(e.get("refreshed_count") or 0), refresh_by_day.get(d, 0))
        e["reactivated_count"] = max(int(e.get("reactivated_count") or 0), react_by_day.get(d, 0))


def generate_dashboard_json(scan_results, scan_timestamp):
    data = load_existing_json()
    now_str = scan_timestamp.strftime("%Y-%m-%d %H:%M:%S")
    today = scan_timestamp.strftime("%Y-%m-%d")
    data["last_scan"] = now_str

    scan_entry = {"timestamp": now_str, "date": today, "profiles": {}}

    for pk, result in scan_results.items():
        cfg = PROFILES[pk]
        if pk not in data["profiles"]:
            data["profiles"][pk] = {
                "label": cfg["label"], "url": cfg["url"],
                "is_category": cfg.get("is_category", False),
                "daily_counts": [], "current_listings": [],
                "archived_listings": [], "price_history": {},
            }

        pd_ = data["profiles"][pk]
        dc = pd_["daily_counts"]

        # Identyfikuj NOWE ogłoszenia (których nie było w poprzednim skanie)
        # UWAGA: nazwa `newly_detected_listings` celowo inna niż `new_listings` niżej —
        # tamta to lista wszystkich przeprocesowanych (stare nazewnictwo, legacy).
        old_ids = set(l["id"] for l in pd_.get("current_listings", []))
        current_ids_new = set(l["listing_id"] for l in result["listings"])
        newly_detected_listings = [l for l in result["listings"] if l["listing_id"] not in old_ids]

        # NEW: Flow per dzień — added = nowo pojawione, removed = znikłe vs poprzedni scan
        # Jeśli to pierwszy scan profilu (brak old_ids) — flow jest None (brak bazy porównania)
        if len(old_ids) == 0 and len(pd_.get("daily_counts", [])) == 0:
            flow_added = None
            flow_removed = None
        else:
            flow_added = len(current_ids_new - old_ids)
            flow_removed = len(old_ids - current_ids_new)

        # Kalkuluj medianę cen TYLKO z nowych ogłoszeń (first_seen == dziś)
        new_prices = [l["price"] for l in newly_detected_listings if l.get("price") is not None and l["price"] > 0]
        if new_prices:
            sorted_prices = sorted(new_prices)
            n = len(sorted_prices)
            if n % 2 == 0:
                median_price = round((sorted_prices[n//2 - 1] + sorted_prices[n//2]) / 2)
            else:
                median_price = sorted_prices[n//2]
        else:
            median_price = None

        # ═══ DATA PROTECTION ═══
        # Chronimy daily_counts i current_listings tylko gdy scan zawiódł.
        # Scenariusze:
        #   1. Scan error (crosscheck=error) → prawdopodobny błąd OLX/network, chroń dane
        #   2. Scan OK ale header_count=None i count=0 → nie można zweryfikować, chroń
        #   3. Scan OK, header=0, count=0 → profil PRAWDZIWIE pusty, archiwizuj normalnie
        #   4. Scan OK, header=N, count=N → normalna sytuacja
        #   5. Scan pobrał < 50% ogłoszeń z nagłówka (np. 50 z 650) → częściowy scrape, chroń dane
        crosscheck = result.get("crosscheck", "")
        header_count = result.get("header_count")
        is_scraper_error = (
            crosscheck == "error"
            or (result["count"] == 0 and header_count is None)
            or is_header_shortfall(result)
        )
        current_listings_count = len(pd_.get("current_listings", []))

        # BUGFIX „ciche gubienie ogłoszeń" (rotacja wyników OLX): ogłoszenia nieobecne
        # w skanie weryfikujemy JUŻ TUTAJ — przed policzeniem flow i przed archiwizacją.
        # Potwierdzone jako wciąż aktywne trafiają do carried_ids: nie liczą się jako
        # `removed` i pozostają w current_listings (patrz pętla archiwizacji niżej),
        # zamiast znikać bez śladu i wracać w kolejnym skanie jako "nowe" z wyzerowaną
        # historią (refresh/reaktywacje/ceny).
        carried_ids = set()
        verified_any = False
        if not is_scraper_error:
            for old_l in pd_.get("current_listings", []):
                if old_l["id"] in current_ids_new:
                    continue
                if not old_l.get("url"):
                    continue
                # Pauza między weryfikacyjnymi GET-ami — seria szybkich requestów
                # może sprowokować blokadę OLX i zafałszować kolejne sprawdzenia.
                if verified_any:
                    time.sleep(0.7)
                verified_any = True
                if verify_listing_active(old_l["url"]):
                    log.info(f"[{pk}] Ogłoszenie nieobecne w skanie, ale aktywne na OLX — zachowuję: {old_l['id']}")
                    carried_ids.add(old_l["id"])
        if flow_removed is not None and carried_ids:
            # Ogłoszenia potwierdzone jako aktywne NIE zniknęły — nie są "removed"
            flow_removed -= len(carried_ids)

        # Guard na daily_counts: pomijamy tylko gdy mamy znane dane a scan zawiódł
        skip_daily_update = is_scraper_error and current_listings_count > 0

        if skip_daily_update:
            log.warning(f"[{pk}] Skipping daily_counts update - scraper error detected (crosscheck={crosscheck}, header={header_count}, current_listings={current_listings_count})")
        else:
            # Build price distribution snapshot for this scan
            def build_price_distribution(listings):
                prices = sorted([l["price"] for l in listings if l.get("price") and l["price"] > 0])
                if not prices:
                    return []
                mn, mx = prices[0], prices[-1]
                if mn == mx:
                    return [{"from": mn, "to": mx + 1, "count": len(prices)}]
                raw = (mx - mn) / 14
                mag = 10 ** max(0, int(len(str(int(max(raw, 1)))) - 1))
                step = next((f * mag for f in [1, 2, 2.5, 5, 10] if f * mag >= raw), 10 * mag)
                start = (mn // step) * step
                buckets = []
                s = start
                while s <= mx:
                    cnt = sum(1 for p in prices if s <= p < s + step)
                    buckets.append({"from": int(s), "to": int(s + step), "count": cnt})
                    s += step
                if prices[-1] >= s - step:
                    extra = sum(1 for p in prices if p >= s)
                    if extra: buckets[-1]["count"] += extra
                while len(buckets) > 1 and buckets[-1]["count"] == 0: buckets.pop()
                while len(buckets) > 1 and buckets[0]["count"] == 0:  buckets.pop(0)
                return buckets

            price_dist = build_price_distribution(result["listings"])

            today_entry = next((d for d in dc if d["date"] == today), None)
            if today_entry:
                if result["count"] >= today_entry["count"]:
                    today_entry["count"] = result["count"]
                    today_entry["timestamp"] = now_str
                    today_entry["median_price"] = median_price

                    # NEW: Update promoted stats
                    total = result["count"]
                    promoted_count = sum(1 for l in result["listings"] if l.get("is_promoted"))
                    promoted_pct = round(promoted_count / total * 100, 1) if total > 0 else 0

                    promo_breakdown = {}
                    for l in result["listings"]:
                        if l.get("is_promoted"):
                            ptype = l.get("promotion_type", 'unknown')
                            promo_breakdown[ptype] = promo_breakdown.get(ptype, 0) + 1

                    today_entry["promoted_count"] = promoted_count
                    today_entry["promoted_percentage"] = promoted_pct
                    today_entry["promotion_breakdown"] = promo_breakdown
                    today_entry["price_distribution"] = price_dist

                    # NEW: Flow (przybyło/zniknęło) — update przy kolejnym scanie dnia
                    # Akumulujemy: added z obecnego scanu + dotychczasowa suma added dnia
                    # Analogicznie removed. Ostatecznie i tak zapisujemy kumulację wszystkich ruchów w ciągu dnia.
                    if flow_added is not None:
                        prev_added = today_entry.get("added") or 0
                        prev_removed = today_entry.get("removed") or 0
                        today_entry["added"] = prev_added + flow_added
                        today_entry["removed"] = prev_removed + flow_removed
                    else:
                        today_entry.setdefault("added", None)
                        today_entry.setdefault("removed", None)

                    # refreshed_count/reactivated_count przelicza recompute_daily_refresh_reactivation()
                    # na końcu przetwarzania profilu (projekcja z historii ogłoszeń)

                    # Przelicz change względem wczoraj, nie poprzedniej wartości dzisiejszej
                    yesterday_entry = dc[-2] if len(dc) >= 2 else None
                    if yesterday_entry:
                        today_entry["change"] = result["count"] - yesterday_entry["count"]
            else:
                prev_c = dc[-1]["count"] if dc else None
                ch = result["count"] - prev_c if prev_c is not None else 0

                # NEW: Calculate promoted stats
                total = result["count"]
                promoted_count = sum(1 for l in result["listings"] if l.get("is_promoted"))
                promoted_pct = round(promoted_count / total * 100, 1) if total > 0 else 0

                # Count by promotion type
                promo_breakdown = {}
                for l in result["listings"]:
                    if l.get("is_promoted"):
                        ptype = l.get("promotion_type", 'unknown')
                        promo_breakdown[ptype] = promo_breakdown.get(ptype, 0) + 1

                # refreshed_count/reactivated_count przelicza recompute_daily_refresh_reactivation()
                # na końcu przetwarzania profilu (projekcja z historii ogłoszeń)

                dc.append({
                    "date": today,
                    "count": result["count"],
                    "change": ch,
                    "timestamp": now_str,
                    "median_price": median_price,
                    # NEW: Promoted metrics
                    "promoted_count": promoted_count,
                    "promoted_percentage": promoted_pct,
                    "promotion_breakdown": promo_breakdown,
                    # Price distribution snapshot
                    "price_distribution": price_dist,
                    # Wartości początkowe — przelicza recompute_daily_refresh_reactivation()
                    "refreshed_count": 0,
                    "reactivated_count": 0,
                    # NEW: Flow (przybyło/zniknęło)
                    "added": flow_added,
                    "removed": flow_removed,
                })

        if len(dc) > 90:
            pd_["daily_counts"] = dc[-90:]

        current_ids = set()
        new_listings = []
        for listing in result["listings"]:
            # Try to get dates directly from listing (JSON scraping provides them)
            # Fall back to parsing date_text for DOM-scraped listings
            pub = listing.get("published")
            ref = listing.get("refreshed")
            if not pub and not ref:
                pub, ref = parse_date_text(listing.get("date_text", ""))
            
            nl = {
                "id": listing["listing_id"], "title": listing["title"],
                "price": listing["price"], "price_text": listing.get("price_text", ""),
                "url": listing["url"], "published": pub, "refreshed": ref,
                "last_refresh_timestamp": listing.get("last_refresh_timestamp"),
                "date_text": listing.get("date_text", ""),
                "image_url": listing.get("image_url", ""),
                "first_seen": now_str, "last_seen": now_str,
                # NEW: Promoted fields
                "is_promoted": listing.get("is_promoted", False),
                "promotion_type": listing.get("promotion_type"),
                # NEW: Refresh count
                "refresh_count": 0,
            }
            new_listings.append(nl)
            current_ids.add(listing["listing_id"])

        old_map = {l["id"]: l for l in pd_.get("current_listings", [])}
        archived_map = {l["id"]: l for l in pd_.get("archived_listings", [])}

        for nl in new_listings:
            lid = nl["id"]
            if lid in old_map:
                old = old_map[lid]
                nl["first_seen"] = old.get("first_seen", now_str)
                # Przenieś flagę reaktywacji jeśli była
                if old.get("reactivated"):
                    nl["reactivated"] = True
                    nl["reactivation_history"] = old.get("reactivation_history", [])
                    nl["reactivation_count"] = len(nl["reactivation_history"])
                
                # REFRESH COUNT & HISTORY TRACKING
                # Skopiuj poprzedni licznik i historię
                nl["refresh_count"] = old.get("refresh_count", 0)
                nl["refresh_history"] = old.get("refresh_history", [])

                # Event odświeżenia = ZMIANA DATY `refreshed` (YYYY-MM-DD) na nowszą
                # LUB pojawienie się daty, gdy wcześniej jej nie było (None -> data).
                # OLX pokazuje na stronie tylko datę ("Odświeżono dnia 12 maja 2026"),
                # więc to jest źródło prawdy. Max 1 event/dzień per ogłoszenie.
                # NIE używamy last_refresh_timestamp jako triggera — w danych category
                # zdarzało się, że ten sam dzień miał różne timestampy (np. przez
                # zmianę strefy lub retry) i generowało to fałszywe duplikaty.
                # Wymóg: porównanie YYYY-MM-DD, deduplikacja po `refreshed_at` (dacie).
                #
                # Liczymy każdą zmianę daty (też dla ogłoszeń promowanych) —
                # jeśli OLX pokazuje nowszą datę odświeżenia, to jest realne odświeżenie.
                old_refreshed = old.get("refreshed")
                new_refreshed = nl.get("refreshed")

                # Zachowaj starą datę `refreshed`, gdy nowy skan jej nie ma (np. karta
                # bez "Odświeżono" / nieudane parsowanie) — analogicznie do `published`.
                # Bez tego data regresowała do None i kolejne odświeżenie nie miało
                # punktu odniesienia (event przepadał).
                if old_refreshed and not new_refreshed:
                    nl["refreshed"] = old_refreshed
                    if not nl.get("last_refresh_timestamp") and old.get("last_refresh_timestamp"):
                        nl["last_refresh_timestamp"] = old["last_refresh_timestamp"]

                is_new_refresh = bool(
                    new_refreshed
                    and (not old_refreshed or new_refreshed > old_refreshed)
                )

                if is_new_refresh:
                    # Deduplikacja per dzień: max 1 wpis w historii dla danej daty.
                    already_counted = any(
                        h.get("refreshed_at") == new_refreshed
                        for h in nl["refresh_history"]
                    )
                    if not already_counted:
                        nl["refresh_count"] += 1
                        nl["refresh_history"].append({
                            "refreshed_at": new_refreshed,
                            "detected_at": now_str,
                            "old_date": old_refreshed,
                        })
                        log.info(f"  [REFRESHED] {lid}: '{nl['title'][:50]}' - odświeżeń: {nl['refresh_count']}")
            elif lid in archived_map:
                # Ogłoszenie wróciło z archiwum — to REAKTYWACJA
                old_archived = archived_map[lid]
                nl["first_seen"] = old_archived.get("first_seen", now_str)
                nl["reactivated"] = True
                history = old_archived.get("reactivation_history", [])
                history.append({
                    "active_from": old_archived.get("first_seen"),
                    "active_to": old_archived.get("archived_date"),
                    "reactivated_at": now_str,
                })
                nl["reactivation_history"] = history
                nl["reactivation_count"] = len(history)
                # Zachowaj refresh_count i refresh_history z archiwum
                nl["refresh_count"] = old_archived.get("refresh_count", 0)
                nl["refresh_history"] = old_archived.get("refresh_history", [])

                # REFRESH DETECTION przy reaktywacji:
                # Jeśli wraca z nowszą datą `refreshed` niż miało w archiwum
                # (lub z datą, gdy w archiwum jej nie było — None -> data),
                # to OLX odświeżył je w trakcie nieaktywności — liczymy event.
                old_archived_refreshed = old_archived.get("refreshed")
                new_refreshed_reactivated = nl.get("refreshed")
                if old_archived_refreshed and not new_refreshed_reactivated:
                    nl["refreshed"] = old_archived_refreshed
                    if not nl.get("last_refresh_timestamp") and old_archived.get("last_refresh_timestamp"):
                        nl["last_refresh_timestamp"] = old_archived["last_refresh_timestamp"]
                if (
                    new_refreshed_reactivated
                    and (not old_archived_refreshed or new_refreshed_reactivated > old_archived_refreshed)
                ):
                    already_counted = any(
                        h.get("refreshed_at") == new_refreshed_reactivated
                        for h in nl["refresh_history"]
                    )
                    if not already_counted:
                        nl["refresh_count"] += 1
                        nl["refresh_history"].append({
                            "refreshed_at": new_refreshed_reactivated,
                            "detected_at": now_str,
                            "old_date": old_archived_refreshed,
                            "during_reactivation": True,
                        })
                        log.info(f"  [REFRESHED-REACT] {lid}: refresh wykryty przy reaktywacji")

                # Usuń z archiwum
                pd_["archived_listings"] = [a for a in pd_["archived_listings"] if a["id"] != lid]
                log.info(f"  [REACTIVATED] {lid}: '{nl['title'][:50]}'")
            if lid in old_map:
                old = old_map[lid]
                
                # LOGIKA NAJSTARSZEJ DATY PUBLIKACJI:
                # Priorytet: published (z OLX) > first_seen (fallback)
                # Zachowaj najstarszą published jeśli była wcześniej zescanowana
                old_published = old.get("published")
                new_published = nl.get("published")
                
                if old_published and new_published:
                    # Wybierz starszą datę między starą a nową published
                    nl["published"] = min(old_published, new_published)
                elif old_published and not new_published:
                    # Zachowaj starą published jeśli nowa nie ma (np. odświeżone)
                    nl["published"] = old_published
                # else: zostaw nową published (jeśli jest) lub None
                
                # Znajdź pierwszą cenę (z first_seen lub najwcześniejszą z price_history)
                first_price = old.get("price")  # Domyślnie wczorajsza cena
                
                # Jeśli jest price_history, weź najstarszą cenę
                if lid in pd_.get("price_history", {}) and pd_["price_history"][lid]:
                    first_price = pd_["price_history"][lid][0]["old_price"]
                
                new_price = nl.get("price")
                old_price = old.get("price")
                
                # Dodaj do historii jeśli cena się zmieniła
                if old_price is not None and new_price is not None and old_price != new_price:
                    if lid not in pd_["price_history"]:
                        pd_["price_history"][lid] = []
                    pd_["price_history"][lid].append({
                        "date": now_str, "old_price": old_price,
                        "new_price": new_price, "change": new_price - old_price,
                    })
                
                # Oblicz price_change względem PIERWSZEJ ceny
                if first_price is not None and new_price is not None:
                    nl["first_price"] = first_price
                    nl["price_change"] = new_price - first_price
                    if nl["price_change"] != 0:
                        nl["previous_price"] = first_price  # Dla kompatybilności
                
                # ═══ PROMOTION TRACKING ═══
                # Track promotion periods: start date, end date, type, consecutive days
                if "promotion_history" not in pd_:
                    pd_["promotion_history"] = {}

                if lid not in pd_["promotion_history"]:
                    pd_["promotion_history"][lid] = []

                old_is_promoted = old.get("is_promoted", False)
                new_is_promoted = nl.get("is_promoted", False)

                # Preserve old promotion data
                nl["promotion_history"] = old.get("promotion_history", [])
                nl["promoted_days_current"] = old.get("promoted_days_current", 0)
                nl["promoted_sessions_count"] = old.get("promoted_sessions_count", 0)

                if new_is_promoted and not old_is_promoted:
                    # STARTED promotion today
                    nl["promotion_started_at"] = today
                    nl["promoted_days_current"] = 1
                    nl["promoted_sessions_count"] = old.get("promoted_sessions_count", 0) + 1
                    nl["_promotion_last_day"] = today
                    log.info(f"  [PROMO START] {lid}: Session #{nl['promoted_sessions_count']}")

                elif new_is_promoted and old_is_promoted:
                    # CONTINUING promotion — inkrementuj tylko raz na dzień
                    nl["promotion_started_at"] = old.get("promotion_started_at", today)
                    last_day = old.get("_promotion_last_day")  # None = pole jeszcze nie istniało
                    if last_day is None:
                        # Pierwsze zetknięcie z guardem — nie inkrementuj, tylko zainicjuj
                        nl["promoted_days_current"] = old.get("promoted_days_current", 1)
                        nl["_promotion_last_day"] = today
                    elif last_day != today:
                        nl["promoted_days_current"] = old.get("promoted_days_current", 0) + 1
                        nl["_promotion_last_day"] = today
                    else:
                        nl["promoted_days_current"] = old.get("promoted_days_current", 0)
                        nl["_promotion_last_day"] = last_day
                    nl["promoted_sessions_count"] = old.get("promoted_sessions_count", 0)

                elif not new_is_promoted and old_is_promoted:
                    # ENDED promotion today
                    promo_start = old.get("promotion_started_at", today)
                    days = old.get("promoted_days_current", 1)

                    # Save to history
                    nl["promotion_history"].append({
                        "start_date": promo_start,
                        "end_date": today,
                        "days": days,
                        "promotion_type": old.get("promotion_type", "unknown"),
                        "session_number": old.get("promoted_sessions_count", 0)
                    })

                    # Reset current counters
                    nl["promoted_days_current"] = 0
                    nl["promoted_sessions_count"] = old.get("promoted_sessions_count", 0)
                    nl.pop("promotion_started_at", None)
                    nl.pop("_promotion_last_day", None)

                    # Save to profile-level history
                    pd_["promotion_history"][lid] = nl["promotion_history"]

                    log.info(f"  [PROMO END] {lid}: Lasted {days} days, session #{nl['promotion_history'][-1]['session_number']}")

            else:
                # ═══ PROMOTION TRACKING DLA NOWYCH / REAKTYWOWANYCH ═══
                # Gdy ogłoszenie pojawia się po raz pierwszy (lub wraca z archiwum)
                # i JEST PROMOWANE — licz dzień 1 zamiast pomijać
                if nl.get("is_promoted"):
                    nl["promotion_started_at"] = today
                    nl["promoted_days_current"] = 1
                    nl["_promotion_last_day"] = today
                    archived_sessions = archived_map.get(lid, {}).get("promoted_sessions_count", 0) if lid in archived_map else 0
                    nl["promoted_sessions_count"] = archived_sessions + 1
                    nl["promotion_history"] = archived_map.get(lid, {}).get("promotion_history", []) if lid in archived_map else []
                    log.info(f"  [PROMO START NEW] {lid}: Session #{nl['promoted_sessions_count']}")
                else:
                    if lid in archived_map:
                        nl["promoted_days_current"] = 0
                        nl["promoted_sessions_count"] = archived_map[lid].get("promoted_sessions_count", 0)
                        nl["promotion_history"] = archived_map[lid].get("promotion_history", [])
                    else:
                        nl["promoted_days_current"] = 0
                        nl["promoted_sessions_count"] = 0
                        nl["promotion_history"] = []

        # CRITICAL: Chroń przed archiwizacją gdy scraper ma błąd (OLX blocking, network, itp.).
        # Gdy profil prawdziwie jest pusty (crosscheck=passed, header=0), archiwizuj normalnie
        # — użytkownik mógł usunąć wszystkie swoje ogłoszenia.
        if not is_scraper_error:
            for old_l in pd_.get("current_listings", []):
                if old_l["id"] not in current_ids:
                    # Weryfikacja (verify_listing_active) odbyła się wyżej, przed flow.
                    # Ogłoszenie potwierdzone jako wciąż aktywne (rotacja wyników OLX)
                    # NIE jest archiwizowane — przenosimy je do nowej listy
                    # current_listings z licznikiem `missed_scans`, żeby zachować
                    # jego historię (refresh/reaktywacje/ceny) między skanami.
                    # Licznik resetuje się sam, gdy ogłoszenie wróci do skanu
                    # (świeży rekord nie kopiuje pola missed_scans).
                    if old_l["id"] in carried_ids:
                        log.info(f"[{pk}] Pominięto archiwizację (ogłoszenie nadal aktywne na OLX) — zachowuję w current_listings: {old_l['id']}")
                        old_l["last_seen"] = now_str
                        old_l["missed_scans"] = int(old_l.get("missed_scans") or 0) + 1
                        new_listings.append(old_l)
                        current_ids.add(old_l["id"])
                        continue
                    if old_l.get("url"):
                        log.info(f"[{pk}] Potwierdzono nieaktywność — archiwizuję: {old_l['id']}")
                    else:
                        log.warning(f"[{pk}] Brak URL dla ogłoszenia {old_l['id']} — archiwizuję bez weryfikacji")

                    old_l["archived_date"] = now_str

                    # Zamknij otwarty okres reaktywacji (jeśli istnieje, czyli ogłoszenie było reaktywowane)
                    # Ostatni wpis w reactivation_history ma reactivated_at ale może nie mieć active_to_closed.
                    # Dodajemy pole active_to_current = archived_date, żeby zachować pełny okres aktywności.
                    r_hist = old_l.get("reactivation_history", [])
                    if r_hist:
                        last = r_hist[-1]
                        # Jeśli ostatni wpis nie ma zamkniętego bieżącego okresu — zamknij go.
                        if "active_to_current" not in last:
                            last["active_to_current"] = now_str

                    # Zachowaj reactivation_count (liczba reaktywacji = długość historii)
                    old_l["reactivation_count"] = len(r_hist)

                    # Upewnij się że refresh_count i refresh_history istnieją
                    if "refresh_history" not in old_l:
                        old_l["refresh_history"] = []
                    if "refresh_count" not in old_l:
                        old_l["refresh_count"] = len(old_l["refresh_history"])

                    pd_["archived_listings"].append(old_l)

            # Archiwum nieograniczone — paginacja po stronie dashboardu

            # Aktualizuj current_listings TYLKO gdy scan był poprawny (count > 0)
            pd_["current_listings"] = new_listings

            # Przelicz dzienne liczniki odświeżeń/reaktywacji we WSZYSTKICH wpisach
            # daily_counts jako projekcję historii ogłoszeń (current + archived).
            # Event wykryty z opóźnieniem (odświeżenie po godzinie skanu, widoczne
            # dopiero nazajutrz z wczorajszą datą) trafia dzięki temu do właściwego
            # dnia wstecz, zamiast przepadać — patrz recompute_daily_refresh_reactivation.
            recompute_daily_refresh_reactivation(pd_)
        else:
            log.warning(f"[{pk}] Skipping archiving AND current_listings update - scraper error detected (crosscheck={crosscheck}, header={header_count})")
            # Zachowaj stare current_listings - nie nadpisuj pustą listą!
        scan_entry["profiles"][pk] = {
            "count": result["count"],
            "crosscheck": result.get("crosscheck", ""),
            "added": flow_added,
            "removed": flow_removed,
        }

    data["scan_history"].append(scan_entry)
    if len(data["scan_history"]) > 90:
        data["scan_history"] = data["scan_history"][-90:]

    # Stabilna serializacja: deterministyczna kolejność list ogłoszeń + kluczy.
    # Dashboard sortuje po stronie klienta, więc kolejność w pliku jest dla niego
    # nieistotna — a stała kolejność drastycznie zmniejsza diff w gicie (mniej churnu).
    for pk, pd_ in data.get("profiles", {}).items():
        if isinstance(pd_.get("current_listings"), list):
            pd_["current_listings"].sort(key=lambda l: str(l.get("id", "")))
        if isinstance(pd_.get("archived_listings"), list):
            pd_["archived_listings"].sort(key=lambda l: str(l.get("id", "")))

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)
    log.info(f"Dashboard JSON saved: {JSON_PATH}")


# ─── API JSON Generation ────────────────────────────────────────────────────

def generate_api_json(scan_results, scan_timestamp, duration_seconds):
    """Generate API status and history JSON files for mobile app (Android)."""
    os.makedirs(API_DIR, exist_ok=True)

    now_iso = scan_timestamp.strftime("%Y-%m-%dT%H:%M:%SZ")
    today_str = scan_timestamp.strftime("%Y-%m-%d")

    total_listings = sum(r["count"] for r in scan_results.values())
    profiles_scanned = len(scan_results)

    # ── Nowe ogłoszenia i zmiany cen per profil ──────────────────────────────
    existing_data = load_existing_json()

    profiles_status = {}
    total_new = 0
    total_removed = 0
    total_price_changes = 0
    error_profiles = []
    alerts = []

    for pk, result in scan_results.items():
        crosscheck = result.get("crosscheck", "unknown")
        count = result.get("count", 0)
        header_count = result.get("header_count")

        # Błąd = jawny "error", 0 ogłoszeń z niepasującym crosscheck,
        # albo częściowy scrape (pobrano < 50% ogłoszeń z nagłówka strony)
        shortfall = is_header_shortfall(result)
        is_error = (
            crosscheck == "error"
            or (count == 0 and crosscheck not in (
                "passed", "passed_retry", "consistent", "best_of_two"
            ))
            or shortfall
        )

        # Przybyło/zniknęło — odczyt z świeżo zapisanego daily_counts (dzisiejszy wpis).
        # generate_dashboard_json() liczy added/removed per profil i zapisuje JSON
        # PRZED tą funkcją, więc bierzemy gotowe wartości zamiast liczyć ponownie.
        prof_data = existing_data.get("profiles", {}).get(pk, {})
        added_today = None
        removed_today = None
        prev_count = None
        for dc_entry in reversed(prof_data.get("daily_counts", [])):
            entry_date = dc_entry.get("date", "")
            if entry_date == today_str and added_today is None:
                added_today = dc_entry.get("added")
                removed_today = dc_entry.get("removed")
            elif entry_date < today_str:
                prev_count = dc_entry.get("count")
                break

        # ── Alerty anomalii ──────────────────────────────────────────────────
        # 1. Częściowy scrape: pobrano drastycznie mniej niż deklaruje nagłówek OLX.
        if shortfall:
            alerts.append({
                "profile": pk,
                "type": "header_shortfall",
                "severity": "critical",
                "message": (f"⚠️ POWAŻNY BŁĄD SKANU: profil „{PROFILES[pk]['label']}” — "
                            f"pobrano tylko {count} z {header_count} ogłoszeń deklarowanych "
                            f"przez OLX. Dane profilu NIE zostały zaktualizowane (ochrona)."),
                "count": count,
                "header_count": header_count,
            })
        # 2. Masowe zniknięcie: ubyło >= 30% wczorajszego stanu profilu (i >= 10 szt.).
        #    Łapie przypadki, gdzie dane JUŻ zostały zmodyfikowane (np. masowa archiwizacja).
        if (removed_today and prev_count
                and removed_today >= MASS_REMOVAL_MIN
                and removed_today >= MASS_REMOVAL_RATIO * prev_count):
            pct = round(100 * removed_today / prev_count)
            alerts.append({
                "profile": pk,
                "type": "mass_removal",
                "severity": "critical",
                "message": (f"⚠️ POWAŻNA ANOMALIA: z profilu „{PROFILES[pk]['label']}” "
                            f"zniknęło {removed_today} z {prev_count} ogłoszeń ({pct}%) "
                            f"w ciągu doby — możliwy błąd skanu/blokada OLX."),
                "removed": removed_today,
                "previous_count": prev_count,
                "count": count,
            })
        # 3. Ogłoszenia "wiszące": od wielu skanów nieobecne w wynikach, a wciąż
        #    uznawane za aktywne. Może znaczyć, że OLX zmienił komunikat o nieaktualności
        #    i verify_listing_active() przestała wykrywać martwe ogłoszenia.
        stale = [l for l in prof_data.get("current_listings", [])
                 if (l.get("missed_scans") or 0) >= STALE_MISSED_SCANS_MIN]
        if stale:
            max_missed = max(l["missed_scans"] for l in stale)
            # Lista podejrzanych ogłoszeń (od najdłużej nieobecnych) — żeby alert był
            # działający: można otworzyć URL i w kilka sekund sprawdzić na oko, czy
            # ogłoszenie faktycznie żyje, czy verify_listing_active() daje false positive.
            stale_items = [
                {
                    "id": l["id"],
                    "url": l.get("url"),
                    "title": l.get("title"),
                    "missed_scans": l["missed_scans"],
                }
                for l in sorted(stale, key=lambda x: -x["missed_scans"])[:10]
            ]
            alerts.append({
                "profile": pk,
                "type": "stale_listings",
                "severity": "warning",
                "message": (f"⚠️ UWAGA: profil „{PROFILES[pk]['label']}” — {len(stale)} "
                            f"ogłoszeń nieobecnych w wynikach od >= {STALE_MISSED_SCANS_MIN} "
                            f"skanów (max {max_missed}), a wciąż uznawanych za aktywne. "
                            f"Możliwa zmiana komunikatu OLX o nieaktualności — sprawdź "
                            f"verify_listing_active() (linki w polu 'stale_items')."),
                "stale_count": len(stale),
                "max_missed_scans": max_missed,
                "stale_items": stale_items,
            })

        new_today = added_today or 0

        # Zmiany cen: wpisy w price_history z dzisiaj
        price_changes_today = 0
        for ph_entries in prof_data.get("price_history", {}).values():
            for entry in ph_entries:
                if entry.get("date", "").startswith(today_str):
                    price_changes_today += 1

        total_new += new_today
        total_removed += (removed_today or 0)
        total_price_changes += price_changes_today

        profile_entry = {
            "label": PROFILES[pk]["label"],
            "count": count,
            "added": added_today,
            "removed": removed_today,
            "new_listings": new_today,
            "price_changes": price_changes_today,
            "crosscheck": crosscheck,
            "crosscheck_detail": result.get("crosscheck_detail", ""),
            "duration_seconds": result.get("duration_seconds"),
            "ok": not is_error,
        }

        if is_error:
            if shortfall:
                detail = (f"Częściowy scrape: pobrano {count} z {header_count} ogłoszeń "
                          f"({result.get('crosscheck_detail', '')})")
            else:
                detail = result.get("crosscheck_detail", "Nieznany błąd podczas skanowania")
            profile_entry["error"] = detail
            error_profiles.append(pk)
        else:
            profile_entry["error"] = None

        profiles_status[pk] = profile_entry

    # ── Globalny status ───────────────────────────────────────────────────────
    error_count = len(error_profiles)
    if error_count == 0 and not alerts:
        status = "success"
        message = f"Skan {profiles_scanned} profili zakończony pomyślnie"
    elif error_count == 0:
        status = "warning"
        alert_profiles = sorted({a["profile"] for a in alerts})
        message = (f"Skan zakończony, ale wykryto poważne anomalie w: "
                   f"{', '.join(alert_profiles)} — sprawdź pole 'alerts'")
    elif error_count < profiles_scanned:
        status = "partial_failure"
        message = f"Skan częściowy — błędy w: {', '.join(error_profiles)}"
    else:
        status = "failure"
        message = "Skan nieudany — wszystkie profile zwróciły błędy"

    for alert in alerts:
        log.warning(f"[ALERT] {alert['message']}")

    # ── Następny scan: 07:00 UTC ──────────────────────────────────────────────
    next_scan = scan_timestamp.replace(hour=7, minute=0, second=0, microsecond=0)
    if scan_timestamp.hour >= 7:
        next_scan += timedelta(days=1)
    next_scan_iso = next_scan.strftime("%Y-%m-%dT%H:%M:%SZ")
    seconds_to_next = max(0, int((next_scan - scan_timestamp).total_seconds()))

    # ── status.json ───────────────────────────────────────────────────────────
    status_data = {
        "status": status,
        "message": message,
        "alerts": alerts,
        "lastScan": {
            "timestamp": now_iso,
            "duration_seconds": duration_seconds,
            "profiles_scanned": profiles_scanned,
            "total_listings": total_listings,
            "added": total_new,
            "removed": total_removed,
            "new_listings": total_new,
            "price_changes": total_price_changes,
            "errors": error_profiles,
        },
        "nextScan": {
            "scheduled": next_scan_iso,
            "in_seconds": seconds_to_next,
        },
        "profiles": profiles_status,
    }

    with open(API_STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(status_data, f, ensure_ascii=False, indent=2)
    log.info(f"API status.json saved: {API_STATUS_PATH}")

    # ── history.json ──────────────────────────────────────────────────────────
    history_data = {"scans": []}
    if os.path.exists(API_HISTORY_PATH):
        try:
            with open(API_HISTORY_PATH, "r", encoding="utf-8") as f:
                history_data = json.load(f)
        except (json.JSONDecodeError, IOError):
            pass

    history_entry = {
        "timestamp": now_iso,
        "date": today_str,
        "status": status,
        "message": message,
        "alerts": alerts,
        "duration_seconds": duration_seconds,
        "total_listings": total_listings,
        "added": total_new,
        "removed": total_removed,
        "new_listings": total_new,
        "price_changes": total_price_changes,
        "profiles_scanned": profiles_scanned,
        "errors": error_profiles,
        "profiles": {
            pk: {
                "label": v["label"],
                "count": v["count"],
                "added": v["added"],
                "removed": v["removed"],
                "new_listings": v["new_listings"],
                "price_changes": v["price_changes"],
                "crosscheck": v["crosscheck"],
                "ok": v["ok"],
                "error": v.get("error"),
            }
            for pk, v in profiles_status.items()
        },
    }

    scans = history_data.get("scans", [])
    scans.append(history_entry)
    scans = scans[-3:]  # API trzyma 3 ostatnie scany

    history_data = {
        "last_updated": now_iso,
        "scans": scans,
        # Shortcut dla aplikacji: 3 najnowsze od razu (od najnowszego)
        "recent": list(reversed(scans)),
    }

    with open(API_HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(history_data, f, ensure_ascii=False, indent=2)
    log.info(f"API history.json saved: {API_HISTORY_PATH}")


# ─── Main ────────────────────────────────────────────────────────────────────

def run_scan():
    # Ensure data directory exists
    os.makedirs(DATA_DIR, exist_ok=True)
    
    ts = datetime.now()
    start_time = time.time()
    
    log.info(f"{'='*60}")
    log.info(f"SZPERACZ OLX — Scan started {ts.strftime('%Y-%m-%d %H:%M:%S')}")
    log.info(f"{'='*60}")

    # All profiles scraped in one Playwright browser session
    results = scrape_with_playwright_all(PROFILES)

    # Odrzuć ogłoszenia z ceną >= 10x średnia w profilu (błędne dane / nie-pokój)
    filter_price_outliers(results)

    duration_seconds = int(time.time() - start_time)
    
    # Najpierw JSON (pełny stan per-ogłoszenie), potem append-only ledger trendu.
    # Excel NIE jest już zapisywany do repo — generowany na żądanie (build_excel_from_data)
    # przy raporcie tygodniowym. Zamrożony backup: data/archive/.
    generate_dashboard_json(results, ts)
    append_history(results, ts)
    generate_api_json(results, ts, duration_seconds)
    generate_trend_full()

    log.info(f"{'='*60}")
    log.info(f"SZPERACZ OLX — Scan completed {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info(f"Duration: {duration_seconds} seconds")
    log.info(f"{'='*60}")
    return results


if __name__ == "__main__":
    run_scan()
