#!/usr/bin/env python3
"""
SZPERACZ OLX — Rebuild liczników odświeżeń i reaktywacji dla ARCHIWUM + wykresu.

STRATEGIA (v2):
1. refresh_history odtwarzamy z DWÓCH sygnałów (bierzemy "unię"):
   (a) Kolumna "Liczba odświeżeń" w Excelu — każdy WZROST wartości w czasie per ID
       = N eventów odświeżenia (gdzie N = różnica). Reset/spadek = IGNORUJEMY
       (to był bug w scraperze w pewnym momencie albo OLX zwraca 0).
       Działa dla WSZYSTKICH profili (user + category).
   (b) Kolumna "Data odświeżenia" — każda ZMIANA daty na nowszą = event.
       Działa tylko dla category ('wszystkie_pokoje').

   Łączymy eventy z (a) i (b), deduplikujemy w obrębie tego samego dnia scanu.

2. reactivation_history — luka >=REACTIVATION_GAP_DAYS scanów bez obecności ID,
   potem powrót.

3. Rebuild daily_counts[*].refreshed_count / reactivated_count z rzeczywistych zdarzeń.

Idempotentny — uruchomienie wielokrotne daje ten sam rezultat.
"""

import json
import os
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
EXCEL_PATH = os.path.join(DATA_DIR, "szperacz_olx.xlsx")
JSON_PATH = os.path.join(DATA_DIR, "dashboard_data.json")

PROFILES = [
    "wszystkie_pokoje",
    "pokojewlublinie",
    "poqui",
    "artymiuk",
    "dawny_patron",
    "mzuri",
    "villahome",
]

REACTIVATION_GAP_DAYS = 2


# ─── I/O ────────────────────────────────────────────────────────────────────

def load_dashboard_json():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_dashboard_json(data):
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _normalize_date(val):
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None


def _normalize_time(val):
    if val is None:
        return "00:00"
    if hasattr(val, "strftime"):
        return val.strftime("%H:%M")
    return str(val).strip() or "00:00"


def _int_or_none(val):
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


# ─── Excel parsing ──────────────────────────────────────────────────────────

def parse_excel_per_listing():
    """
    Returns per-profile, per-listing chronological timeline:
      {
        "profile_key": {
          "listing_id": [
             {"scan_date", "scan_time", "refreshed_date", "refresh_count_col"},
             ...
          ]
        }
      }
    """
    print("=" * 70)
    print("🔍 PARSOWANIE EXCEL — per-listing timeline")
    print("=" * 70)

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Brak pliku Excel: {EXCEL_PATH}")
        return {}

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    result = {}

    for profile_key in PROFILES:
        sheet_name = profile_key[:31]
        if sheet_name not in wb.sheetnames:
            print(f"   ⚠️  [{profile_key}] brak arkusza — pomijam")
            continue

        ws = wb[sheet_name]
        data = defaultdict(list)

        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        header = {cell.value: idx for idx, cell in enumerate(header_row)}

        required = ["Data scanu", "Godzina", "Tytuł", "ID ogłoszenia"]
        missing = [c for c in required if c not in header]
        if missing:
            print(f"   ⚠️  [{profile_key}] brak kluczowych kolumn {missing} — pomijam")
            continue

        col_scan_date = header["Data scanu"]
        col_scan_time = header["Godzina"]
        col_title = header["Tytuł"]
        col_listing_id = header["ID ogłoszenia"]
        col_refreshed = header.get("Data odświeżenia")
        col_refresh_count = header.get("Liczba odświeżeń")

        indices = [col_scan_date, col_scan_time, col_title, col_listing_id]
        if col_refreshed is not None:
            indices.append(col_refreshed)
        if col_refresh_count is not None:
            indices.append(col_refresh_count)
        max_col_needed = max(indices)

        for row in ws.iter_rows(min_row=2):
            row_values = [cell.value for cell in row]
            if len(row_values) <= max_col_needed:
                continue

            scan_date = _normalize_date(row_values[col_scan_date])
            scan_time = _normalize_time(row_values[col_scan_time])
            title = row_values[col_title]
            listing_id = row_values[col_listing_id]

            if not listing_id or not title:
                continue

            refreshed_date = _normalize_date(row_values[col_refreshed]) if col_refreshed is not None else None
            refresh_count_col = _int_or_none(row_values[col_refresh_count]) if col_refresh_count is not None else None

            data[str(listing_id)].append({
                "scan_date": scan_date,
                "scan_time": scan_time,
                "refreshed_date": refreshed_date,
                "refresh_count_col": refresh_count_col,
            })

        for lid in data:
            data[lid].sort(key=lambda e: (e["scan_date"] or "", e["scan_time"] or ""))

        result[profile_key] = dict(data)
        print(f"   ✓ [{profile_key}] {len(data)} unikalnych ID")

    wb.close()
    return result


# ─── Refresh history reconstruction ─────────────────────────────────────────

def build_refresh_events(entries):
    """
    Buduje refresh_history[] na podstawie dwóch sygnałów:

    (a) Wzrosty 'Liczba odświeżeń' — każdy wzrost z X → Y (Y>X) = (Y-X) eventów.
        Resety/spadki ignorujemy, ale high-water mark aktualizujemy tylko w górę.

    (b) Zmiany 'Data odświeżenia' na nowszą datę.

    Dedup po scan_date: w obrębie tego samego dnia scanu liczymy event RAZ
    (bo nie wiemy czy wzrost count i zmiana daty to ten sam event, czy dwa).

    Returns: (history_list, count)
    """
    events = []

    # ── (a) Increments of 'Liczba odświeżeń' (high-water mark logic) ──
    hwm = None  # high-water mark
    for e in entries:
        cur = e["refresh_count_col"]
        if cur is None:
            continue
        if hwm is None:
            hwm = cur
            continue
        if cur > hwm:
            delta = cur - hwm
            for _ in range(delta):
                events.append({
                    "detected_at": f"{e['scan_date']} {e['scan_time']}",
                    "scan_date": e["scan_date"],
                    "refreshed_at": e["scan_date"],  # placeholder — dokładna data unknown
                    "source": "count_increment",
                })
            hwm = cur
        # cur <= hwm: ignoruj (reset lub brak zmiany)

    # ── (b) Zmiany 'Data odświeżenia' ──
    prev_refreshed = None
    date_change_events = []
    for e in entries:
        cur = e["refreshed_date"]
        if cur and prev_refreshed and cur > prev_refreshed:
            date_change_events.append({
                "detected_at": f"{e['scan_date']} {e['scan_time']}",
                "scan_date": e["scan_date"],
                "refreshed_at": cur,
                "old_date": prev_refreshed,
                "source": "date_change",
            })
        if cur:
            prev_refreshed = cur

    # ── Merge z dedup po scan_date ──
    # Zasada: jeśli w ten sam scan_date mamy wzrost count I zmianę daty,
    # licz jako 1 event (preferując dokładniejsze info z date_change).
    events_by_date = defaultdict(list)
    for ev in events + date_change_events:
        events_by_date[ev["scan_date"]].append(ev)

    merged = []
    for scan_date, group in events_by_date.items():
        count_events = [e for e in group if e["source"] == "count_increment"]
        date_events = [e for e in group if e["source"] == "date_change"]

        if count_events and date_events:
            # Ta sama data: weź MAX(count_events, date_events) — bo to pewnie ten sam event
            # (jeśli count wzrósł o 2, ale mamy tylko 1 date_change w tym dniu, liczymy 2)
            total = max(len(count_events), len(date_events))
            # Dla pierwszych N użyj date_change (ma dokładniejsze info)
            for i in range(total):
                if i < len(date_events):
                    merged.append(date_events[i])
                else:
                    merged.append(count_events[i - len(date_events)])
        else:
            merged.extend(group)

    merged.sort(key=lambda x: x["detected_at"])

    # Usuń internal 'source' i 'scan_date'
    for h in merged:
        h.pop("source", None)
        h.pop("scan_date", None)

    return merged, len(merged)


def build_reactivation_events(entries, all_scan_dates, gap_days=REACTIVATION_GAP_DAYS):
    """
    Reaktywacja = ID nieobecne na >=gap_days kolejnych scanach, potem wraca.
    """
    presence_days = sorted(set(e["scan_date"] for e in entries if e["scan_date"]))
    if not presence_days:
        return [], 0, []

    scan_dates_list = sorted(all_scan_dates)
    scan_date_to_idx = {d: i for i, d in enumerate(scan_dates_list)}

    periods = []
    current_start = presence_days[0]
    current_end = presence_days[0]

    for i in range(1, len(presence_days)):
        prev = presence_days[i - 1]
        cur = presence_days[i]

        idx_prev = scan_date_to_idx.get(prev)
        idx_cur = scan_date_to_idx.get(cur)

        if idx_prev is not None and idx_cur is not None:
            gap_scans = idx_cur - idx_prev - 1
        else:
            d_prev = datetime.strptime(prev, "%Y-%m-%d")
            d_cur = datetime.strptime(cur, "%Y-%m-%d")
            gap_scans = (d_cur - d_prev).days - 1

        if gap_scans >= gap_days:
            periods.append((current_start, current_end))
            current_start = cur
        current_end = cur

    periods.append((current_start, current_end))

    history = []
    for i in range(1, len(periods)):
        prev_start, prev_end = periods[i - 1]
        cur_start, _ = periods[i]
        history.append({
            "active_from": f"{prev_start} 00:00:00",
            "active_to": f"{prev_end} 23:59:59",
            "reactivated_at": f"{cur_start} 00:00:00",
        })

    return history, len(history), periods


# ─── Main rebuild ───────────────────────────────────────────────────────────

def rebuild_all():
    data = load_dashboard_json()
    excel_data = parse_excel_per_listing()

    scan_dates_by_profile = defaultdict(set)
    for pk, pd_ in data.get("profiles", {}).items():
        for dc in pd_.get("daily_counts", []):
            if dc.get("date"):
                scan_dates_by_profile[pk].add(dc["date"])

    print()
    print("=" * 70)
    print("🔨 REKONSTRUKCJA HISTORII z Excela")
    print("=" * 70)

    stats = {
        "listings_processed": 0,
        "refresh_events_total": 0,
        "reactivation_events_total": 0,
        "archived_with_refresh": 0,
        "archived_with_reactivation": 0,
        "active_with_refresh": 0,
        "active_with_reactivation": 0,
    }

    refresh_per_day = defaultdict(lambda: defaultdict(int))
    reactivation_per_day = defaultdict(lambda: defaultdict(int))

    for pk in PROFILES:
        if pk not in data.get("profiles", {}):
            continue
        pd_ = data["profiles"][pk]
        per_listing = excel_data.get(pk, {})
        scan_dates = scan_dates_by_profile.get(pk, set())

        current = pd_.get("current_listings", [])
        archived = pd_.get("archived_listings", [])

        profile_stats = {"r_events": 0, "a_events": 0, "listings_with_r": 0, "listings_with_a": 0}

        for listings, is_archived in [(current, False), (archived, True)]:
            for listing in listings:
                lid = str(listing.get("id") or "")
                if not lid:
                    continue
                stats["listings_processed"] += 1
                entries = per_listing.get(lid, [])

                if not entries:
                    listing.setdefault("refresh_history", [])
                    listing.setdefault("refresh_count", 0)
                    listing.setdefault("reactivation_history", [])
                    listing["reactivation_count"] = len(listing["reactivation_history"])
                    continue

                # Refresh history
                refresh_hist, refresh_cnt = build_refresh_events(entries)
                listing["refresh_history"] = refresh_hist
                listing["refresh_count"] = refresh_cnt

                for ev in refresh_hist:
                    det_date = ev["detected_at"].split(" ")[0]
                    refresh_per_day[pk][det_date] += 1

                # Reactivations
                reactivation_hist, reactivation_cnt, _ = build_reactivation_events(
                    entries, scan_dates, gap_days=REACTIVATION_GAP_DAYS
                )
                listing["reactivation_history"] = reactivation_hist
                listing["reactivation_count"] = reactivation_cnt

                if reactivation_cnt > 0:
                    listing["reactivated"] = True

                for ev in reactivation_hist:
                    ra_date = ev["reactivated_at"].split(" ")[0]
                    reactivation_per_day[pk][ra_date] += 1

                if is_archived and reactivation_hist:
                    last = reactivation_hist[-1]
                    if "active_to_current" not in last:
                        archived_date = listing.get("archived_date")
                        if archived_date:
                            last["active_to_current"] = archived_date

                if is_archived:
                    if refresh_cnt > 0:
                        stats["archived_with_refresh"] += 1
                    if reactivation_cnt > 0:
                        stats["archived_with_reactivation"] += 1
                else:
                    if refresh_cnt > 0:
                        stats["active_with_refresh"] += 1
                    if reactivation_cnt > 0:
                        stats["active_with_reactivation"] += 1

                stats["refresh_events_total"] += refresh_cnt
                stats["reactivation_events_total"] += reactivation_cnt
                profile_stats["r_events"] += refresh_cnt
                profile_stats["a_events"] += reactivation_cnt
                if refresh_cnt > 0:
                    profile_stats["listings_with_r"] += 1
                if reactivation_cnt > 0:
                    profile_stats["listings_with_a"] += 1

        print(f"📊 [{pk}] current={len(current)} archived={len(archived)} "
              f"| listings z refresh: {profile_stats['listings_with_r']}, "
              f"z reakt: {profile_stats['listings_with_a']} "
              f"| eventy: refresh={profile_stats['r_events']}, reakt={profile_stats['a_events']}")

    print()
    print("=" * 70)
    print("🔨 REBUILD daily_counts.refreshed_count / reactivated_count")
    print("=" * 70)

    for pk in PROFILES:
        if pk not in data.get("profiles", {}):
            continue
        pd_ = data["profiles"][pk]
        dc = pd_.get("daily_counts", [])

        updated = 0
        for entry in dc:
            date = entry.get("date")
            if not date:
                continue
            r = refresh_per_day[pk].get(date, 0)
            a = reactivation_per_day[pk].get(date, 0)
            old_r = entry.get("refreshed_count", 0)
            old_a = entry.get("reactivated_count", 0)
            entry["refreshed_count"] = r
            entry["reactivated_count"] = a
            if r != old_r or a != old_a:
                updated += 1

        total_r = sum(refresh_per_day[pk].values())
        total_a = sum(reactivation_per_day[pk].values())
        print(f"   [{pk}] updated {updated} dni | suma refresh={total_r}, reakt={total_a}")

    save_dashboard_json(data)

    print()
    print("=" * 70)
    print("✅ PODSUMOWANIE")
    print("=" * 70)
    print(f"   Ogłoszeń przetworzonych:      {stats['listings_processed']}")
    print(f"   Eventów odświeżeń:            {stats['refresh_events_total']}")
    print(f"   Eventów reaktywacji:          {stats['reactivation_events_total']}")
    print(f"   Aktywnych z refresh_count>0:  {stats['active_with_refresh']}")
    print(f"   Aktywnych z reakt_count>0:    {stats['active_with_reactivation']}")
    print(f"   Archiwum z refresh_count>0:   {stats['archived_with_refresh']}")
    print(f"   Archiwum z reakt_count>0:     {stats['archived_with_reactivation']}")
    print()
    print(f"   Zapisano: {JSON_PATH}")


if __name__ == "__main__":
    rebuild_all()
