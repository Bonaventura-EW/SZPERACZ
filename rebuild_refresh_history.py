#!/usr/bin/env python3
"""
SZPERACZ OLX — Rebuild refresh_history from Excel data.
Przechodzi przez wszystkie wiersze w Excel dla każdego ogłoszenia
i buduje pełną historię zmian daty 'refreshed'.
"""

import json
import os
from collections import defaultdict
from openpyxl import load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
EXCEL_PATH = os.path.join(DATA_DIR, "szperacz_olx.xlsx")
JSON_PATH = os.path.join(DATA_DIR, "dashboard_data.json")

PROFILES = [
    "wszystkie_pokoje",
    "pokojewlublinie",
    "poqui",
    "artymiuk",
    "dawny_patron",
    "mzuri",
    "villahome",
]


def load_dashboard_json():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_dashboard_json(data):
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def extract_refresh_history_from_excel():
    """
    Przechodzi przez wszystkie profile w Excel i ekstrahuje historię odświeżeń.
    
    Struktura:
    {
        "profile_key": {
            "listing_id": [
                {"scan_date": "2026-03-01", "scan_time": "09:15", "refreshed": "2026-02-28"},
                {"scan_date": "2026-03-05", "scan_time": "09:20", "refreshed": "2026-03-03"},
                ...
            ]
        }
    }
    """
    print("=" * 70)
    print("🔍 ODTWARZANIE REFRESH_HISTORY Z EXCEL")
    print("=" * 70)
    
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    result = {}
    
    for profile_key in PROFILES:
        sheet_name = profile_key[:31]  # Excel truncates to 31 chars
        
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  [{profile_key}] Brak arkusza '{sheet_name}', pomijam")
            continue
        
        ws = wb[sheet_name]
        profile_data = defaultdict(list)
        
        print(f"\n📊 [{profile_key}] Parsowanie arkusza...")
        
        # Autodetekcja kolumn z nagłówka (odporna na zmiany layoutu)
        header = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}
        
        col_scan_date = header.get("Data scanu")
        col_scan_time = header.get("Godzina")
        col_title = header.get("Tytuł")
        col_refreshed = header.get("Data odświeżenia")
        col_listing_id = header.get("ID ogłoszenia")
        
        # Walidacja — wszystkie kluczowe kolumny muszą istnieć
        missing = [name for name, idx in [
            ("Data scanu", col_scan_date),
            ("Godzina", col_scan_time),
            ("Tytuł", col_title),
            ("Data odświeżenia", col_refreshed),
            ("ID ogłoszenia", col_listing_id),
        ] if idx is None]
        
        if missing:
            print(f"   ⚠️  [{profile_key}] Brak kolumn w nagłówku: {missing}, pomijam")
            continue
        
        print(f"   Kolumny: scan_date={col_scan_date+1}, refreshed={col_refreshed+1}, id={col_listing_id+1}")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_values = [cell.value for cell in row]
            
            max_col_needed = max(col_scan_date, col_scan_time, col_title, col_refreshed, col_listing_id)
            if len(row_values) <= max_col_needed:
                continue
            
            scan_date = row_values[col_scan_date]
            scan_time = row_values[col_scan_time]
            title = row_values[col_title]
            refreshed = row_values[col_refreshed]
            listing_id = row_values[col_listing_id]
            
            # Pomiń puste wiersze i summary rows (brak ID)
            if not listing_id or not title:
                continue
            
            # Konwertuj daty do stringów
            if scan_date:
                scan_date = scan_date.strftime("%Y-%m-%d") if hasattr(scan_date, 'strftime') else str(scan_date)
            if scan_time:
                scan_time = scan_time.strftime("%H:%M") if hasattr(scan_time, 'strftime') else str(scan_time)
            if refreshed:
                refreshed = refreshed.strftime("%Y-%m-%d") if hasattr(refreshed, 'strftime') else str(refreshed)
            
            # Dodaj tylko jeśli mamy wszystkie dane
            if scan_date and listing_id:
                profile_data[listing_id].append({
                    "scan_date": scan_date,
                    "scan_time": scan_time or "00:00",
                    "refreshed": refreshed or None,
                    "title": title,
                })
        
        result[profile_key] = dict(profile_data)
        print(f"   ✓ Znaleziono {len(profile_data)} unikalnych ID")
    
    wb.close()
    return result


def build_refresh_history(excel_data, data):
    """
    Dla każdego ogłoszenia w current_listings i archived_listings:
    - Przeszukaj dane z Excel
    - Znajdź wszystkie zmiany daty 'refreshed'
    - Zbuduj refresh_history[]
    """
    print("\n" + "=" * 70)
    print("🔨 BUDOWANIE REFRESH_HISTORY")
    print("=" * 70)
    
    total_histories_built = 0
    total_refresh_events = 0
    
    for profile_key, profile_data in data["profiles"].items():
        print(f"\n📋 [{profile_key}]")
        
        if profile_key not in excel_data:
            print(f"   ⚠️  Brak danych Excel, pomijam")
            continue
        
        excel_profile = excel_data[profile_key]
        
        # Przetwórz current_listings
        last_scan_ts = data.get("last_scan", "")
        last_scan_date = last_scan_ts.split(" ")[0] if last_scan_ts else ""
        last_scan_time = last_scan_ts.split(" ")[1][:5] if " " in last_scan_ts else "00:00"
        
        for listing in profile_data.get("current_listings", []):
            lid = listing["id"]
            
            # Zbierz skany z Excel (jeśli są)
            scans = list(excel_profile.get(lid, []))
            
            # FALLBACK: jeśli JSON ma 'refreshed' dla tego ogłoszenia i ostatni scan w Excel
            # tego nie pokazuje (rozbieżność Excel vs JSON), dodaj syntetyczny wpis z last_scan
            json_refreshed = listing.get("refreshed")
            if json_refreshed and last_scan_date:
                excel_has_current = any(
                    s.get("refreshed") == json_refreshed for s in scans
                )
                if not excel_has_current:
                    scans.append({
                        "scan_date": last_scan_date,
                        "scan_time": last_scan_time,
                        "refreshed": json_refreshed,
                        "title": listing.get("title", ""),
                    })
            
            if not scans:
                continue
            
            # Posortuj chronologicznie
            scans.sort(key=lambda x: (x["scan_date"], x["scan_time"]))
            
            # Znajdź zmiany refreshed date
            # Logika: każda UNIKALNA data refreshu liczona jest osobno
            # Pierwsze pojawienie się daty (prev_refreshed=None → curr_refreshed=data) też się liczy
            history = []
            prev_refreshed = None
            
            for scan in scans:
                curr_refreshed = scan["refreshed"]
                
                if curr_refreshed and curr_refreshed != prev_refreshed:
                    # Nowa data refreshu (pierwsza lub zmieniona) - dodaj do historii
                    history.append({
                        "refreshed_at": curr_refreshed,
                        "detected_at": f"{scan['scan_date']} {scan['scan_time']}:00",
                        "old_date": prev_refreshed  # None dla pierwszego refreshu
                    })
                
                if curr_refreshed:
                    prev_refreshed = curr_refreshed
            
            # ZAWSZE nadpisz (nie tylko gdy history jest niepuste) —
            # inaczej stary, buggy refresh_count z JSON zostałby zachowany
            listing["refresh_history"] = history
            listing["refresh_count"] = len(history)
            if history:
                total_histories_built += 1
                total_refresh_events += len(history)
                print(f"   ✓ {lid[:8]}... '{listing['title'][:40]}' → {len(history)} odświeżeń")
        
        # Przetwórz archived_listings
        for listing in profile_data.get("archived_listings", []):
            lid = listing["id"]
            
            scans = list(excel_profile.get(lid, []))
            
            # Fallback: jeśli JSON ma 'refreshed' a Excel tego nie pokazuje
            json_refreshed = listing.get("refreshed")
            archived_at = listing.get("archived_date", "")
            if json_refreshed and archived_at:
                archived_date = archived_at.split(" ")[0]
                archived_time = archived_at.split(" ")[1][:5] if " " in archived_at else "00:00"
                excel_has_current = any(
                    s.get("refreshed") == json_refreshed for s in scans
                )
                if not excel_has_current:
                    scans.append({
                        "scan_date": archived_date,
                        "scan_time": archived_time,
                        "refreshed": json_refreshed,
                        "title": listing.get("title", ""),
                    })
            
            if not scans:
                continue
            
            scans.sort(key=lambda x: (x["scan_date"], x["scan_time"]))
            history = []
            prev_refreshed = None
            
            for scan in scans:
                curr_refreshed = scan["refreshed"]
                
                if curr_refreshed and curr_refreshed != prev_refreshed:
                    # Nowa data refreshu (pierwsza lub zmieniona)
                    history.append({
                        "refreshed_at": curr_refreshed,
                        "detected_at": f"{scan['scan_date']} {scan['scan_time']}:00",
                        "old_date": prev_refreshed
                    })
                
                if curr_refreshed:
                    prev_refreshed = curr_refreshed
            
            # ZAWSZE nadpisz (dla spójności z current_listings)
            listing["refresh_history"] = history
            listing["refresh_count"] = len(history)
            if history:
                total_histories_built += 1
                total_refresh_events += len(history)
    
    print("\n" + "=" * 70)
    print("📊 PODSUMOWANIE")
    print("=" * 70)
    print(f"✅ Ogłoszeń z historią: {total_histories_built}")
    print(f"✅ Łącznie wydarzeń odświeżenia: {total_refresh_events}")
    print("=" * 70)


def main():
    # 1. Wczytaj JSON
    data = load_dashboard_json()
    
    # 2. Wyciągnij dane z Excel
    excel_data = extract_refresh_history_from_excel()
    
    # 3. Zbuduj refresh_history
    build_refresh_history(excel_data, data)
    
    # 4. Zapisz JSON
    save_dashboard_json(data)
    print(f"\n💾 Zapisano: {JSON_PATH}")


if __name__ == "__main__":
    main()
